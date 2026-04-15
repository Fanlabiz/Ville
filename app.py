"""
Optimiseur de placement de batiments - Application Streamlit
Compatible iPad Excel (francais) - Deploiement GitHub/Streamlit Cloud
"""

import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import copy

st.set_page_config(page_title="Optimiseur de Ville", layout="wide")
st.title("Optimiseur de placement de batiments")

# ══════════════════════════════════════════════════════
# CONSTANTES COULEURS (format ARGB 8 caracteres)
# ══════════════════════════════════════════════════════
C_ORANGE  = "FFFFA500"
C_GREEN   = "FF90EE90"
C_GRAY    = "FFD3D3D3"
C_BLUE    = "FF4472C4"
C_WHITE   = "FFFFFFFF"
C_BORDX   = "FF808080"
C_GAIN    = "FF006400"
C_LOSS    = "FFCC0000"

def mfill(hex8):
    return PatternFill("solid", fgColor=hex8)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color=C_WHITE)
    cell.fill = mfill(C_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

# ══════════════════════════════════════════════════════
# LECTURE DES DONNEES
# ══════════════════════════════════════════════════════

def read_terrain(ws):
    """Retourne grid[r][c] = 'X'|None, max_r, max_c (0-indexes)."""
    max_r, max_c = ws.max_row, ws.max_column
    grid = [[None] * max_c for _ in range(max_r)]
    for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
        for cell in row:
            if cell.value == "X":
                grid[cell.row - 1][cell.column - 1] = "X"
    return grid, max_r, max_c


def _to_float(val, default=0.0):
    """
    Convertit une valeur en float de facon robuste.
    Gere les formules Excel non evaluees (ex: '=49980/2').
    """
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return default
    if s.startswith("="):
        import re
        expr = s[1:]
        if re.fullmatch(r"[\d\s\+\-\*\/\(\)\.]+", expr):
            try:
                return float(eval(expr))
            except Exception:
                pass
        return default
    try:
        return float(s)
    except (ValueError, TypeError):
        return default


def read_buildings_def(ws):
    """Lit l'onglet Batiments. Retourne liste de dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c else "" for c in rows[0]]
    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        d = dict(zip(header, row))
        result.append({
            "nom":        str(d.get("Nom", "")).strip(),
            "longueur":   int(_to_float(d.get("Longueur"), 1)),
            "largeur":    int(_to_float(d.get("Largeur"), 1)),
            "nombre":     int(_to_float(d.get("Nombre"), 1)),
            "type":       str(d.get("Type", "")).strip(),
            "culture":    _to_float(d.get("Culture")),
            "rayonnement":int(_to_float(d.get("Rayonnement"))),
            "boost25":    _to_float(d.get("Boost 25%")),
            "boost50":    _to_float(d.get("Boost 50%")),
            "boost100":   _to_float(d.get("Boost 100%")),
            "production": str(d.get("Production", "")).strip(),
            "quantite":   _to_float(d.get("Quantite")),
            "priorite":   _to_float(d.get("Priorite"), 0.0),
            "placement":  str(d.get("Placement", "Obligatoire")).strip()
                          if d.get("Placement") else "Obligatoire",
        })
    return result


def read_placed_buildings(ws):
    """
    Lit les batiments places sur le terrain :
    - Cellules fusionnees  -> batiments multi-cases
    - Cellules simples non-X -> batiments 1x1
    Retourne liste de dicts {nom, r, c, rows, cols}.
    """
    placed = []
    merged_set = set()

    for mr in ws.merged_cells.ranges:
        top = ws.cell(mr.min_row, mr.min_col)
        name = str(top.value).strip() if top.value else ""
        if name and name != "X":
            placed.append({
                "nom":  name,
                "r":    mr.min_row - 1,
                "c":    mr.min_col - 1,
                "rows": mr.max_row - mr.min_row + 1,
                "cols": mr.max_col - mr.min_col + 1,
            })
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_set.add((r, c))

    for row in ws.iter_rows():
        for cell in row:
            if (cell.row, cell.column) in merged_set:
                continue
            if cell.value and cell.value != "X":
                placed.append({
                    "nom":  str(cell.value).strip(),
                    "r":    cell.row - 1,
                    "c":    cell.column - 1,
                    "rows": 1,
                    "cols": 1,
                })
    return placed


def enrich(placed, buildings_def):
    """Ajoute les infos catalogue a chaque batiment place."""
    catalog = {b["nom"].strip(): b for b in buildings_def}
    result = []
    for p in placed:
        base = catalog.get(p["nom"].strip(), {
            "type": "Neutre", "culture": 0, "rayonnement": 0,
            "boost25": 0, "boost50": 0, "boost100": 0,
            "production": "Rien", "quantite": 0, "priorite": 0,
            "longueur": p["cols"], "largeur": p["rows"], "nombre": 1,
        })
        result.append({**base, **p})
    return result


# ══════════════════════════════════════════════════════
# MECANIQUE CULTURE / BOOST / SCORE
# ══════════════════════════════════════════════════════

def cells_of(b):
    """Cases occupees par le batiment b."""
    return {(b["r"] + dr, b["c"] + dc)
            for dr in range(b["rows"]) for dc in range(b["cols"])}


def radiation_zone(b):
    """Cases dans la zone de rayonnement du batiment culturel b."""
    ray = b.get("rayonnement", 0)
    r0, c0 = b["r"], b["c"]
    r1, c1 = r0 + b["rows"] - 1, c0 + b["cols"] - 1
    return {
        (r, c)
        for r in range(r0 - ray, r1 + ray + 1)
        for c in range(c0 - ray, c1 + ray + 1)
        if not (r0 <= r <= r1 and c0 <= c <= c1)
    }


def culture_received(producer, culturels):
    """Culture totale recue par un batiment producteur."""
    prod_cells = cells_of(producer)
    return sum(cult["culture"] for cult in culturels
               if prod_cells & radiation_zone(cult))


def boost_level(culture, b):
    """Boost obtenu (0, 25, 50 ou 100)."""
    if b["type"] != "Producteur":
        return 0
    if b["boost100"] and culture >= b["boost100"]:
        return 100
    if b["boost50"] and culture >= b["boost50"]:
        return 50
    if b["boost25"] and culture >= b["boost25"]:
        return 25
    return 0


def score_placement(placed):
    """Score total = somme(boost*priorite) pour les Producteurs (grande priorite = plus important)."""
    culturels = [b for b in placed if b["type"] == "Culturel"]
    total = 0.0
    for b in placed:
        if b["type"] == "Producteur" and b["priorite"] > 0:
            cult = culture_received(b, culturels)
            boost = boost_level(cult, b)
            total += boost * b["priorite"]
    return total


# ══════════════════════════════════════════════════════
# PLACEMENT INITIAL DES BATIMENTS MANQUANTS
# ══════════════════════════════════════════════════════

def place_missing_buildings(placed, buildings_def, terrain_grid, max_r, max_c,
                            n_trials=10, time_budget=20.0, force_complete=True):
    """
    Place les batiments manquants en combinant FFD multi-start et
    un scoring guide par des regles metier :

    Regle 1 (Neutres en bordure) :
      Les batiments Neutres preferent les cases adjacentes aux X,
      pour laisser le coeur du terrain aux batiments qui impactent le score.

    Regle 2 (Culturels orientes vers les producteurs) :
      Un culturel est place a la position ou sa zone de rayonnement contient
      le plus de cases libres ou deja occupees par des producteurs.
      Cases libres = futurs producteurs potentiels.
      Cases neutre/culturel dans la zone = inutiles pour ce calcul.

    Regle 3 (Eviter la fragmentation) :
      Apres placement hypothetique, penaliser les positions qui creent
      des espaces residuels trop petits ou trop etroits (incapables
      d'accueillir un batiment de taille minimale significative).

    Ces regles sont des preferences, pas des obligations.
    Si aucune position ideale n'existe, on prend la moins mauvaise
    pour garantir le placement de tous les batiments.

    Ordre de placement : Producteurs -> Culturels -> Neutres
    (les culturels voient les producteurs deja en place pour la regle 2).

    Retourne (new_placed, n_placed, n_failed).
    """
    import random, time
    from collections import Counter
    from itertools import groupby

    placed_counts = Counter(b["nom"].strip() for b in placed)
    to_place_base = []
    to_place_optional = []  # tentés en dernier uniquement en mode score maximal
    for b_def in buildings_def:
        nom = b_def["nom"].strip()
        is_optional = b_def.get("placement", "Obligatoire").strip().lower() == "optionnel"
        needed = b_def["nombre"] - placed_counts.get(nom, 0)
        for _ in range(needed):
            entry = dict(b_def, nom=nom,
                         rows=b_def["largeur"],
                         cols=b_def["longueur"],
                         is_optional=is_optional)
            if is_optional and not force_complete:
                # Mode "score maximal": les optionnels sont exclus du placement principal.
                # Ils seront tentés après optimisation si de la place reste.
                to_place_optional.append(entry)
            else:
                to_place_base.append(entry)

    if not to_place_base:
        return [dict(b) for b in placed], 0, 0, to_place_optional

    x_grid = make_x_grid(terrain_grid, max_r, max_c)

    # Pre-calculer la distance au X le plus proche pour chaque case interieure
    # via BFS multi-source depuis toutes les cases X simultanement -> O(n²)
    from collections import deque
    dist_to_x = [[999] * max_c for _ in range(max_r)]
    bfs_queue = deque()
    for r in range(max_r):
        for c in range(max_c):
            if terrain_grid[r][c] == "X":
                dist_to_x[r][c] = 0
                bfs_queue.append((r, c))
    while bfs_queue:
        r, c = bfs_queue.popleft()
        for dr, dc in ((-1,0),(1,0),(0,-1),(0,1)):
            nr, nc = r+dr, c+dc
            if 0 <= nr < max_r and 0 <= nc < max_c and dist_to_x[nr][nc] == 999:
                dist_to_x[nr][nc] = dist_to_x[r][c] + 1
                bfs_queue.append((nr, nc))


    def build_occ(placed_list):
        occ = [[False] * max_c for _ in range(max_r)]
        for b in placed_list:
            for dr in range(b["rows"]):
                for dc in range(b["cols"]):
                    rr, cc = b["r"] + dr, b["c"] + dc
                    if 0 <= rr < max_r and 0 <= cc < max_c:
                        occ[rr][cc] = True
        return occ

    def cells_rect(r, c, rows, cols):
        return {(r + dr, c + dc) for dr in range(rows) for dc in range(cols)}

    # ── Regle 2 : score culturel ──
    def score_culturel(r, c, rows, cols, ray, occ, prod_cells_set):
        """
        Compte les cases utiles dans la zone de rayonnement du culturel
        si place en (r,c) avec taille rows×cols et rayonnement ray.
        Utile = libre ET dans le terrain, OU deja occupee par un producteur.
        """
        r0, c0 = r, c
        r1, c1 = r + rows - 1, c + cols - 1
        count = 0
        for rz in range(r0 - ray, r1 + ray + 1):
            for cz in range(c0 - ray, c1 + ray + 1):
                if r0 <= rz <= r1 and c0 <= cz <= c1:
                    continue  # case du batiment lui-meme
                if not (0 <= rz < max_r and 0 <= cz < max_c):
                    continue
                if x_grid[rz][cz]:
                    continue
                if (rz, cz) in prod_cells_set:
                    count += 2   # producteur deja en place : bonus
                elif not occ[rz][cz]:
                    count += 1   # case libre : futur producteur possible
        return count

    # ── Regle 3 : penalite fragmentation ──
    MIN_USEFUL = 3  # taille minimale d'un bloc utile (3 cases dans au moins une dimension)

    def fragmentation_penalty(r, c, rows, cols, occ):
        """
        Apres placement hypothetique en (r,c), verifie le voisinage immediat.
        Compte les zones libres adjacentes qui deviendraient trop petites ou
        trop etroites (< MIN_USEFUL cases dans chaque dimension d'un rectangle
        minimal). On ne scanne que le voisinage elargi pour rester rapide.
        """
        # Marquer temporairement les cases du batiment comme occupees
        temp_occ_cells = cells_rect(r, c, rows, cols)

        penalty = 0
        # Regarder les espaces libres contigus dans le voisinage elargi (rayon 2)
        margin = 2
        visited = set()
        for rn in range(r - margin, r + rows + margin):
            for cn in range(c - margin, c + cols + margin):
                if (rn, cn) in visited or (rn, cn) in temp_occ_cells:
                    continue
                if not (0 <= rn < max_r and 0 <= cn < max_c):
                    continue
                if x_grid[rn][cn] or occ[rn][cn]:
                    continue
                # BFS pour mesurer la taille et les dimensions de la zone libre adjacente
                zone = set()
                queue = [(rn, cn)]
                visited.add((rn, cn))
                while queue:
                    qr, qc = queue.pop()
                    zone.add((qr, qc))
                    for dr, dc in ((-1,0),(1,0),(0,-1),(0,1)):
                        nr, nc = qr+dr, qc+dc
                        if (nr, nc) not in visited and (nr, nc) not in temp_occ_cells:
                            if 0 <= nr < max_r and 0 <= nc < max_c:
                                if not x_grid[nr][nc] and not occ[nr][nc]:
                                    visited.add((nr, nc))
                                    queue.append((nr, nc))
                if not zone:
                    continue
                # Calculer les dimensions du bounding box de la zone
                rs = [zr for zr, zc in zone]
                cs = [zc for zr, zc in zone]
                height = max(rs) - min(rs) + 1
                width  = max(cs) - min(cs) + 1
                # Penaliser si trop petit ou trop etroit
                if len(zone) < MIN_USEFUL or height < 2 or width < 2:
                    penalty += 1
        return penalty

    # ── Fonction de score rapide (regles 1 et 2 seulement) ──
    def fast_score(b, r, c, rows, cols, occ, prod_cells_set):
        """Score rapide sans fragmentation, pour le screening initial."""
        s = 0.0
        if b["type"] == "Neutre":
            avg_dist = sum(dist_to_x[r+dr][c+dc]
                           for dr in range(rows) for dc in range(cols)) / (rows * cols)
            s += 10.0 / (1.0 + avg_dist)
        elif b["type"] == "Culturel":
            ray = b.get("rayonnement", 1)
            s += score_culturel(r, c, rows, cols, ray, occ, prod_cells_set)
        return s

    TOP_N = 10  # nombre de candidats evalues pour la regle 3

    def smart_place_one(ordered, seed):
        """
        Place les batiments dans l'ordre donne.

        Strategie par type :
        - Producteurs : first-fit pur (raster scan). Les regles metier
                        s'appliquent aux Culturels et Neutres, pas aux
                        Producteurs dont la position optimale sera trouvee
                        par l'optimiseur greedy apres le placement initial.
        - Culturels   : scoring regle 2 (cases prod/libres dans rayonnement)
                        + regle 3 (fragmentation, seuil strict <=1).
                        Fallback first-fit si fragmentation trop elevee.
        - Neutres     : scoring regle 1 (proximite des bords X)
                        + regle 3 (fragmentation, seuil genereux <=2).
                        Fallback first-fit si fragmentation trop elevee.
        """
        random.seed(seed)
        result = [dict(b) for b in placed]
        n_ok = n_fail = 0

        for b in ordered:
            occ = build_occ(result)

            # Trouver la premiere position raster valide (fallback absolu)
            first_fit = None
            all_candidates = []
            prod_cells_set = None

            for rows, cols in [(b["rows"], b["cols"]), (b["cols"], b["rows"])]:
                for r in range(max_r):
                    for c in range(max_c):
                        if can_place(r, c, rows, cols, x_grid, occ, max_r, max_c):
                            if first_fit is None:
                                first_fit = (r, c, rows, cols)
                            # Scorer seulement Culturels et Neutres
                            if b["type"] != "Producteur":
                                if b["type"] == "Culturel" and prod_cells_set is None:
                                    prod_cells_set = {
                                        (pb["r"] + dr, pb["c"] + dc)
                                        for pb in result if pb["type"] == "Producteur"
                                        for dr in range(pb["rows"])
                                        for dc in range(pb["cols"])
                                    }
                                s = fast_score(b, r, c, rows, cols, occ,
                                               prod_cells_set or set())
                                all_candidates.append((s, r, c, rows, cols))

            if first_fit is None:
                n_fail += 1
                continue

            if b["type"] == "Producteur" or not all_candidates:
                # Producteur : toujours first-fit
                r, c, rows, cols = first_fit
            else:
                all_candidates.sort(key=lambda x: -x[0])

                # Seuil de fragmentation selon le type
                max_penalty = 1 if b["type"] == "Culturel" else 2

                best_pos   = None
                best_final = None
                for s0, r, c, rows, cols in all_candidates[:TOP_N]:
                    penalty = fragmentation_penalty(r, c, rows, cols, occ)
                    final_s = s0 - 2.0 * penalty
                    if best_final is None or final_s > best_final:
                        best_final = final_s
                        best_pos   = (r, c, rows, cols)

                r, c, rows, cols = best_pos
                # Si fragmentation trop elevee -> first_fit
                if fragmentation_penalty(r, c, rows, cols, occ) > max_penalty:
                    r, c, rows, cols = first_fit

            result.append({**b, "r": r, "c": c, "rows": rows, "cols": cols})
            n_ok += 1

        return result, n_ok, n_fail

    def quick_score(placed_list, max_inner=2):
        """Score direct du placement sans passes greedy.
        Avec le placement en séries (prod + culturels groupés),
        le score initial reflète déjà la qualité du placement.
        Les passes greedy cachaient les différences entre trials."""
        return score_placement(placed_list)

    # Tri des bâtiments pour ffd_pure :
    # - Non-Neutres (Producteurs + Culturels) : taille décroissante ENSEMBLE.
    #   En les mélangeant dans le même flux, un Culturel de 3x4 se place
    #   naturellement à côté d'un Producteur de 3x4 → meilleure couverture.
    # - Neutres : taille décroissante, placés en bordure (regle 1).
    priority = {"Producteur": 0, "Neutre": 1, "Culturel": 2}
    sorted_base = sorted(to_place_base,
                         key=lambda b: (priority.get(b["type"], 9), -(b["rows"] * b["cols"])))

    sorted_ffd = sorted(to_place_base, key=lambda b: -(b["rows"] * b["cols"]))
    # BFS multi-source depuis toutes les cases X
    from collections import deque as _deque
    _dist = [[9999] * max_c for _ in range(max_r)]
    _bfs = _deque()
    for _r in range(max_r):
        for _c in range(max_c):
            if terrain_grid[_r][_c] == "X":
                _dist[_r][_c] = 0; _bfs.append((_r, _c))
    while _bfs:
        _r, _c = _bfs.popleft()
        for _dr, _dc in ((-1,0),(1,0),(0,-1),(0,1)):
            _nr, _nc = _r+_dr, _c+_dc
            if 0<=_nr<max_r and 0<=_nc<max_c and _dist[_nr][_nc]==9999:
                _dist[_nr][_nc] = _dist[_r][_c]+1; _bfs.append((_nr,_nc))
    # Cases internes triees par distance croissante au bord (bordure d'abord)
    cells_border_first = sorted(
        [(_r, _c) for _r in range(max_r) for _c in range(max_c) if not x_grid[_r][_c]],
        key=lambda rc: _dist[rc[0]][rc[1]]
    )
    # Cases internes triees par distance croissante au CENTRE (scan centripète)
    # Utilisé pour les casernes haute priorité: elles se placent au centre du terrain
    # et peuvent ainsi être entourées de culturels de tous côtés.
    _ctr_r, _ctr_c = max_r / 2.0, max_c / 2.0
    cells_center_first = sorted(
        [(_r, _c) for _r in range(max_r) for _c in range(max_c) if not x_grid[_r][_c]],
        key=lambda rc: abs(rc[0] - _ctr_r) + abs(rc[1] - _ctr_c)
    )

    # Listes séparées pour ffd_pure :
    # - non_neutres_ffd : Producteurs + Culturels avec tri prioritaire :
    #   1. Producteurs à haute priorité (casernes prio>=10) en tête, par taille décroissante
    #   2. Culturels forts (culture*rayonnement élevé) immédiatement après pour couvrir les casernes
    #   3. Reste des bâtiments par taille décroissante
    #   Ainsi les casernes se groupent en haut-gauche et les culturels forts les entourent.
    # - neutres_ffd     : Neutres triés par taille décroissante, placés en bordure
    max_prio = max((b["priorite"] for b in to_place_base if b["type"] == "Producteur"), default=1)
    prio_threshold = max_prio * 0.8  # considère "haute priorité" les 20% supérieurs

    # non_neutres_ffd : Producteurs + Culturels mélangés par taille décroissante.
    # Ce mélange naturel garantit 0 échecs et une bonne intrication spatiale.
    # Le biais dans ffd_pure oriente chaque Culturel vers les casernes en priorité.
    # La greedy_pass triée par priorité concentre ensuite les culturels sur les casernes.
    def _nn_sort_key(b):
        # Utilisé pour le groupby dans les trials suivants
        return -(b["rows"] * b["cols"])

    non_neutres_ffd = sorted(
        [b for b in to_place_base if b["type"] != "Neutre"],
        key=_nn_sort_key
    )
    neutres_ffd = sorted(
        [b for b in to_place_base if b["type"] == "Neutre"],
        key=lambda b: -(b["rows"] * b["cols"])
    )

    def ffd_pure(nn_ordered, n_ordered):
        """
        FFD en séries par priorité décroissante :
        1. Neutres en bordure (garantit 0 echecs).
        2. Pour chaque producteur haute priorité (par prio desc, seuil100 asc) :
             a. Placer le producteur au barycentre des producteurs haute prio déjà placés.
             b. Immédiatement après, placer les culturels non encore placés qui
                peuvent le couvrir, du plus fort au plus faible, jusqu'à ce que
                le producteur atteigne son seuil 100% (ou qu'il n'y ait plus
                de culturels disponibles pouvant le couvrir).
           -> Les culturels forts se placent autour de chaque caserne dès sa pose,
              avant que les positions proches soient prises par autre chose.
        3. Reste (autres producteurs + culturels résiduels) en first-fit raster.
        """
        result = [dict(b) for b in placed]
        n_ok = n_fail = 0
        placed_prods = []
        placed_top_prods = []

        # Séparer les bâtiments
        top_prods_all = sorted(
            [b for b in nn_ordered
             if b["type"] == "Producteur" and b["priorite"] >= prio_threshold],
            key=lambda b: (-(b["rows"] * b["cols"]), b.get("boost100", 9999))
        )
        # cults_pool sera défini après Phase 1 (dépend de cults_faibles_ids)
        autres = [b for b in nn_ordered
                  if b["type"] == "Producteur" and b["priorite"] < prio_threshold]
        cults_placed_ids = set()  # ids des culturels déjà placés en série

        # ── Phase 1 : Neutres + producteurs bas-prio + culturels faibles en bordure ──
        # Placer en bordure :
        # - Les neutres
        # - Les producteurs bas-prio (fermes, maisons, etc.) pour garantir leur placement
        # - Les culturels "faibles" (culture*rayonnement faible) qui n'ont pas besoin d'être
        #   au centre du cluster, afin de laisser de l'espace pour les puissants
        # Cela garantit un placement 0-échec ET laisse le centre pour le cluster optimal.
        prods_bas_prio = [b for b in nn_ordered
                          if b["type"] == "Producteur" and b["priorite"] < prio_threshold]
        prods_bas_prio_sorted = sorted(prods_bas_prio, key=lambda b: -(b["rows"] * b["cols"]))
        prods_bas_prio_ids = {id(b) for b in prods_bas_prio}

        # Calculer la puissance de chaque culturel (culture * rayonnement)
        _cults_in_nn = [b for b in nn_ordered if b["type"] == "Culturel"]
        _cult_powers = sorted([b["culture"] * max(b.get("rayonnement", 1), 1)
                                for b in _cults_in_nn])
        # Seuil: les 40% les plus faibles vont en bordure
        _power_threshold = _cult_powers[int(len(_cult_powers) * 0.55)] if _cult_powers else 0
        cults_faibles = [b for b in _cults_in_nn
                         if b["culture"] * max(b.get("rayonnement", 1), 1) <= _power_threshold]
        cults_faibles_sorted = sorted(cults_faibles, key=lambda b: -(b["rows"] * b["cols"]))
        cults_faibles_ids = {id(b) for b in cults_faibles}

        # Pool de culturels pour le cluster: exclure ceux déjà en bordure
        cults_pool = [b for b in nn_ordered
                      if b["type"] == "Culturel" and id(b) not in cults_faibles_ids]

        # Séparer obligatoires et optionnels pour placement en bordure
        _border_obligatoires = [b for b in n_ordered + prods_bas_prio_sorted + cults_faibles_sorted
                                  if not b.get("is_optional", False)]
        _border_optionnels   = [b for b in n_ordered + prods_bas_prio_sorted + cults_faibles_sorted
                                  if b.get("is_optional", False)]

        for b in _border_obligatoires + _border_optionnels:
            occ = build_occ(result)
            ok = False
            for r, c in cells_border_first:
                for rows, cols in [(b["rows"], b["cols"]), (b["cols"], b["rows"])]:
                    if can_place(r, c, rows, cols, x_grid, occ, max_r, max_c):
                        result.append({**b, "r": r, "c": c, "rows": rows, "cols": cols})
                        n_ok += 1; ok = True; break
                if ok: break
            if not ok:
                # Optionnel non placé: pas une erreur bloquante
                if not b.get("is_optional", False):
                    n_fail += 1

        # ── Phase 2 : Séries (producteur haute prio + ses culturels) ──
        # Ordre : PRIORITÉ DÉCROISSANTE => les bâtiments haute prio reçoivent les culturels en premier.
        # Si priorité égale: seuil boost100 décroissant (plus difficile en premier).
        top_prods_all_sorted = sorted(
            top_prods_all,
            key=lambda b: (-b.get("priorite", 0), -b.get("boost100", 0), -(b["rows"] * b["cols"]))
        )

        for prod in top_prods_all_sorted:

            # 2a. Placer le producteur au centre du terrain
            occ = build_occ(result)
            _tr, _tc = max_r / 2.0, max_c / 2.0
            chosen = None; _best_d = 9999
            for r in range(max_r):
                for c in range(max_c):
                    for rows, cols in [(prod["rows"], prod["cols"]),
                                       (prod["cols"], prod["rows"])]:
                        if can_place(r, c, rows, cols, x_grid, occ, max_r, max_c):
                            d = abs(r+rows/2-_tr)+abs(c+cols/2-_tc)
                            if d < _best_d:
                                _best_d=d; chosen=(r, c, rows, cols)
            if chosen:
                r, c, rows, cols = chosen
                result.append({**prod, "r": r, "c": c, "rows": rows, "cols": cols})
                placed_prod_ref = result[-1]
                placed_prods.append(placed_prod_ref)
                placed_top_prods.append(placed_prod_ref)
                n_ok += 1
            else:
                n_fail += 1
                continue

            # 2b. Placer les culturels AU PLUS PRES du producteur, les plus puissants en premier.
            # On exclut les très petits culturels (surface <= 2: sites 1x1, 2x1) car ils
            # fragmentent le terrain et empêchent les gros bâtiments de Phase 3 de se placer.
            # Exception: si sans eux on n'atteint pas 100%, on les inclut en dernier recours.
            prod_boost100 = placed_prod_ref.get("boost100", 9999)
            # D'abord: culturels de surface > 2 uniquement
            cults_grands = sorted(
                [c for c in cults_pool if id(c) not in cults_placed_ids
                 and c["rows"] * c["cols"] > 2],
                key=lambda c: -(c["culture"] * max(c.get("rayonnement", 1), 1))
            )
            # Ensuite: tous (si les grands ne suffisent pas)
            cults_tous = sorted(
                [c for c in cults_pool if id(c) not in cults_placed_ids],
                key=lambda c: -(c["culture"] * max(c.get("rayonnement", 1), 1))
            )
            # Phase 2b: placer les culturels par ordre de puissance décroissante
            # en privilégiant les grands (surface >= 4) qui ne fragmentent pas.
            # On vise 100% en Phase 2b uniquement avec des grands culturels.
            # Les petits (2x1, 1x1) restent pour Phase 3 (ils remplissent les interstices).
            # Vérifier si les grands suffisent pour atteindre boost100:
            _cult_max_grands = sum(c["culture"] for c in cults_grands)
            _cult_deja = culture_received(placed_prod_ref, [b for b in result if b["type"]=="Culturel"])
            if _cult_deja + _cult_max_grands >= prod_boost100:
                cults_disponibles = cults_grands  # les grands suffisent
            else:
                # Les grands ne suffisent pas: inclure les petits mais en dernier
                cults_disponibles = cults_tous
            prod_cr = placed_prod_ref["r"] + placed_prod_ref["rows"] / 2.0
            prod_cc = placed_prod_ref["c"] + placed_prod_ref["cols"] / 2.0

            for cult in cults_disponibles:
                cults_actuels = [b for b in result if b["type"] == "Culturel"]
                cult_recue = culture_received(placed_prod_ref, cults_actuels)
                if cult_recue >= prod_boost100:
                    break  # deja a 100%

                occ = build_occ(result)
                ray = cult.get("rayonnement", 1)
                prod_cells_ref = cells_of(placed_prod_ref)

                # Trouver la position valide LA PLUS PROCHE qui couvre ce producteur
                best_pos = None
                best_dist_cult = float("inf")

                for rows, cols in [(cult["rows"], cult["cols"]),
                                   (cult["cols"], cult["rows"])]:
                    for r in range(max_r):
                        for c in range(max_c):
                            if not can_place(r, c, rows, cols, x_grid, occ, max_r, max_c):
                                continue
                            r0, c0, r1, c1 = r, c, r+rows-1, c+cols-1
                            zone = {(rz, cz)
                                    for rz in range(r0-ray, r1+ray+1)
                                    for cz in range(c0-ray, c1+ray+1)
                                    if not (r0<=rz<=r1 and c0<=cz<=c1)
                                    and 0<=rz<max_r and 0<=cz<max_c}
                            if prod_cells_ref & zone:
                                d = abs(r + rows/2 - prod_cr) + abs(c + cols/2 - prod_cc)
                                if d < best_dist_cult:
                                    best_dist_cult = d
                                    best_pos = (r, c, rows, cols)

                if best_pos:
                    r, c, rows, cols = best_pos
                    result.append({**cult, "r": r, "c": c,
                                   "rows": rows, "cols": cols})
                    cults_placed_ids.add(id(cult))
                    n_ok += 1

        # ── Phase 3 : Reste (autres producteurs + culturels non placés) ──
        _cults_residuels = [c for c in cults_pool if id(c) not in cults_placed_ids]
        # Exclure les producteurs bas-prio et culturels faibles déjà placés en Phase 1
        autres_restants = [b for b in autres if id(b) not in prods_bas_prio_ids]
        # Les culturels faibles sont déjà placés en Phase 1; ne les mettre dans résidus
        # que s'ils n'ont pas encore été placés (au cas où Phase 1 aurait échoué pour certains)
        _cults_residuels_filtres = [c for c in _cults_residuels
                                    if id(c) not in cults_faibles_ids or
                                    not any(b["nom"] == c["nom"] for b in result
                                            if b["type"] == "Culturel")]
        _tout_reste = _cults_residuels_filtres + autres_restants
        # Tri : producteurs 2x2 (surface=4) en tête, puis BFD (grands en premier)
        # avec culturels avant producteurs à taille égale.
        # Ordre Phase 3:
        # Groupe 0: tous les producteurs restants, taille croissante
        #   -> les petits prods (2x2, 3x3, 3x4) trouvent leurs cases AVANT
        #      que les culturels 1x1/1x2 fragmentent le terrain.
        # Groupe 1: culturels résiduels grands (surface > 2), taille desc
        # Groupe 2: culturels résiduels très petits (surface <= 2), taille asc
        #   -> remplissent les tout petits interstices en dernier.
        # Phase 3 : ordre optimisé pour minimiser la fragmentation.
        # Tous les bâtiments de surface >= 9 (prod ET culturels) par taille asc,
        # puis les bâtiments 4-8 cases (2x2, 2x3...) par taille asc,
        # puis les tout petits culturels (1-3 cases) en dernier.
        # -> Les producteurs et grands culturels trouvent leurs blocs AVANT
        #    que les petits culturels fragmentent le terrain.
        def _phase3_key(b):
            surf = b["rows"] * b["cols"]
            # Producteurs en priorité (ils doivent trouver leurs blocs avant
            # que les culturels résiduels fragmentent le terrain).
            # Puis culturels grands, puis petits culturels en dernier.
            if b["type"] == "Producteur":
                if surf >= 9:
                    return (0, surf)   # grands producteurs en tête
                else:
                    return (1, surf)   # petits producteurs ensuite
            else:  # Culturel
                if surf >= 9:
                    return (2, surf)   # grands culturels
                elif surf >= 4:
                    return (3, surf)   # culturels moyens
                else:
                    return (4, surf)   # petits culturels en dernier (1x1, 1x2)
        reste = sorted(_tout_reste, key=_phase3_key)

        for b in reste:
            occ = build_occ(result)
            if b["type"] == "Culturel":
                # Biais : couvrir un producteur haute prio > tout autre producteur
                first_fit = None; chosen_top = None; chosen_any = None
                for rows, cols in [(b["rows"], b["cols"]), (b["cols"], b["rows"])]:
                    if chosen_top: break
                    for r in range(max_r):
                        if chosen_top: break
                        for c in range(max_c):
                            if can_place(r, c, rows, cols, x_grid, occ,
                                         max_r, max_c):
                                if first_fit is None:
                                    first_fit = (r, c, rows, cols)
                                ray = b.get("rayonnement", 1)
                                r0, c0, r1, c1 = r, c, r+rows-1, c+cols-1
                                zone = {(rz, cz)
                                        for rz in range(r0-ray, r1+ray+1)
                                        for cz in range(c0-ray, c1+ray+1)
                                        if not (r0<=rz<=r1 and c0<=cz<=c1)
                                        and 0<=rz<max_r and 0<=cz<max_c}
                                if placed_top_prods and any(
                                        cells_of(p) & zone
                                        for p in placed_top_prods):
                                    chosen_top = (r, c, rows, cols); break
                                elif chosen_any is None and placed_prods and any(
                                        cells_of(p) & zone
                                        for p in placed_prods):
                                    chosen_any = (r, c, rows, cols)
                chosen = chosen_top or chosen_any or first_fit
            else:
                chosen = None
                for rows, cols in [(b["rows"], b["cols"]), (b["cols"], b["rows"])]:
                    if chosen: break
                    for r in range(max_r):
                        if chosen: break
                        for c in range(max_c):
                            if can_place(r, c, rows, cols, x_grid, occ,
                                         max_r, max_c):
                                chosen = (r, c, rows, cols); break

            if chosen:
                r, c, rows, cols = chosen
                result.append({**b, "r": r, "c": c, "rows": rows, "cols": cols})
                if b["type"] == "Producteur":
                    placed_prods.append(result[-1])
                n_ok += 1
            else:
                n_fail += 1

        return result, n_ok, n_fail

    best_placed      = None
    best_screen_score = -1
    best_n_placed    = 0
    best_n_failed    = len(to_place_base)
    t_start = time.time()

    # ── Trials : FFD avec mélange aléatoire des non-Neutres de même taille ──
    # Le trial 0 utilise l'ordre pur (taille décroissante).
    # Les trials suivants mélangent à l'intérieur de chaque groupe de taille,
    # ce qui peut produire de meilleurs mélanges Producteurs/Culturels.
    for trial in range(n_trials + 1):
        if trial > 0 and time.time() - t_start > time_budget:
            break

        random.seed(trial * 17 + 3)
        if trial == 0:
            nn_ordered = non_neutres_ffd  # ordre pur
            n_ordered  = neutres_ffd
        else:
            groups = []
            for _, g in groupby(non_neutres_ffd,
                                key=lambda b: -(b["rows"] * b["cols"])):
                grp = list(g); random.shuffle(grp); groups.append(grp)
            nn_ordered = [b for g in groups for b in g]
            # Mélanger aussi les Neutres (même taille) pour varier
            n_groups = []
            for _, g in groupby(neutres_ffd, key=lambda b: b["rows"] * b["cols"]):
                grp = list(g); random.shuffle(grp); n_groups.append(grp)
            n_ordered = [b for g in n_groups for b in g]

        result, n_ok, n_fail = ffd_pure(nn_ordered, n_ordered)

        # Compter les échecs par catégorie :
        # - Producteurs haute priorité (casernes) manquants : INACCEPTABLE
        # - Autres non-Neutres (culturels, prods faibles) manquants : récupérables
        _placed_noms = Counter(b["nom"].strip() for b in result)
        n_fail_top = sum(
            max(0, b_def["nombre"] - _placed_noms.get(b_def["nom"].strip(), 0))
            for b_def in buildings_def
            if b_def["type"] == "Producteur"
            and b_def.get("priorite", 0) >= prio_threshold
            and b_def.get("placement", "Obligatoire").strip().lower() != "optionnel"
        )

        # Priorité absolue : 0 échecs sur les casernes OBLIGATOIRES haute prio.
        # Les Optionnels peuvent être ignorés même en mode force_complete=False.
        if n_fail_top > 0:
            continue  # jamais acceptable

        s = quick_score(result)

        if (n_fail < best_n_failed or
                (n_fail == best_n_failed and s > best_screen_score)):
            best_placed       = result
            best_screen_score = s
            best_n_placed     = n_ok
            best_n_failed     = n_fail

    if best_placed is None:
        # Fallback : placement simple sans scoring
        result = [dict(b) for b in placed]
        n_ok = n_fail = 0
        for b in sorted_base:
            occ = build_occ(result)
            ok = False
            for rows, cols in [(b["rows"], b["cols"]), (b["cols"], b["rows"])]:
                if ok: break
                for r in range(max_r):
                    if ok: break
                    for c in range(max_c):
                        if can_place(r, c, rows, cols, x_grid, occ, max_r, max_c):
                            result.append({**b, "r": r, "c": c, "rows": rows, "cols": cols})
                            n_ok += 1; ok = True; break
            if not ok:
                n_fail += 1
        best_placed, best_n_placed, best_n_failed = result, n_ok, n_fail

    # ── Post-processing : placer les batiments manquants en FFD pur ──
    # Recupere tous les bâtiments non places (Neutres, culturels,
    # petits producteurs) que le trial en séries n'a pas pu caser
    # faute de place au moment de leur traitement.
    placed_counts_final = Counter(b["nom"].strip() for b in best_placed)
    missing_to_place = []
    for b_def in buildings_def:
        nom = b_def["nom"].strip()
        missing = b_def["nombre"] - placed_counts_final.get(nom, 0)
        for _ in range(missing):
            missing_to_place.append(dict(b_def, nom=nom,
                                         rows=b_def["largeur"],
                                         cols=b_def["longueur"]))

    if missing_to_place:
        # Trier par taille décroissante pour minimiser la fragmentation
        missing_to_place.sort(key=lambda b: -(b["rows"] * b["cols"]))
        n_recovered = n_still_fail = 0

        for b in missing_to_place:
            occ = build_occ(best_placed)
            ok = False
            for rows, cols in [(b["rows"], b["cols"]), (b["cols"], b["rows"])]:
                if ok: break
                for r in range(max_r):
                    if ok: break
                    for c in range(max_c):
                        if can_place(r, c, rows, cols, x_grid, occ, max_r, max_c):
                            best_placed.append({**b, "r": r, "c": c,
                                                "rows": rows, "cols": cols})
                            n_recovered += 1; ok = True; break
            if not ok:
                n_still_fail += 1

        best_n_placed += n_recovered
        best_n_failed  = n_still_fail

    # ── Push algorithm : déplacer des bâtiments légers pour libérer de la place ──
    # Si des bâtiments restent non placés, chercher des blocs cibles bloqués
    # uniquement par des bâtiments de faible priorité et petite surface,
    # les déplacer vers d'autres positions libres, et placer le manquant.
    _max_prio_push = max((b.get("priorite", 0) for b in best_placed
                          if b["type"] == "Producteur"), default=1)
    _prio_thresh_push = _max_prio_push * 0.8

    if best_n_failed > 0:
        _placed_counts_push = Counter(b["nom"].strip() for b in best_placed)
        _still_missing_push = []
        for b_def in buildings_def:
            nom = b_def["nom"].strip()
            missing = b_def["nombre"] - _placed_counts_push.get(nom, 0)
            for _ in range(missing):
                _still_missing_push.append(dict(b_def, nom=nom,
                                                rows=b_def["largeur"],
                                                cols=b_def["longueur"]))

        if _still_missing_push:
            result_push = [dict(b) for b in best_placed]  # copie mutable

            def _find_best_block_and_push(result_p, target_rows, target_cols):
                """
                Trouve le meilleur bloc target_rows x target_cols potentiel
                bloqué UNIQUEMENT par des bâtiments de faible prio (< _prio_thresh_push)
                et déplace ces bâtiments pour libérer le bloc.
                Retourne True si succès, False sinon.
                """
                _cell_map = {}
                for b in result_p:
                    for dr in range(b["rows"]):
                        for dc in range(b["cols"]):
                            _cell_map[(b["r"]+dr, b["c"]+dc)] = b

                occ_p = build_occ(result_p)

                # Trouver le bloc optimal: minimum de surface des blockers
                best_block = None; best_block_score = float("inf")
                for r in range(max_r - target_rows + 1):
                    for c in range(max_c - target_cols + 1):
                        if x_grid[r][c]: continue
                        _blockers = {}; _valid = True
                        for dr in range(target_rows):
                            for dc in range(target_cols):
                                rr, cc = r+dr, c+dc
                                if rr >= max_r or cc >= max_c or x_grid[rr][cc]:
                                    _valid = False; break
                                if occ_p[rr][cc]:
                                    b = _cell_map.get((rr, cc))
                                    if b: _blockers[id(b)] = b
                            if not _valid: break
                        if not _valid: continue
                        if not _blockers:
                            return True, r, c  # bloc libre direct!
                        _blocker_list = list(_blockers.values())
                        _max_p = max(b.get("priorite", 0) for b in _blocker_list)
                        # Ne pousser que les neutres et petits culturels (surface<=2)
                        # Ne jamais pousser des producteurs (même faible prio) car
                        # ils peuvent être couverts par des culturels -> casser les boosts.
                        _blocker_types = {b['type'] for b in _blocker_list}
                        _all_pushable = all(
                            b['type'] == 'Neutre' or
                            (b['type'] == 'Culturel' and b['rows'] * b['cols'] <= 2)
                            for b in _blocker_list
                        )
                        if not _all_pushable: continue  # bloquer non-pushable
                        _score = sum(b["rows"]*b["cols"] for b in _blocker_list)
                        if _score < best_block_score:
                            best_block_score = _score
                            best_block = (r, c, _blocker_list)

                if best_block is None:
                    return False, -1, -1

                r, c, _blocker_list = best_block
                # Déplacer les blockers vers d'autres positions libres
                _reserve = {(r+dr, c+dc) for dr in range(target_rows) for dc in range(target_cols)}
                for _blocker in _blocker_list:
                    # Retirer ce blocker du résultat
                    result_p[:] = [b for b in result_p if b is not _blocker]
                    occ_p = build_occ(result_p)
                    # Trouver une nouvelle position qui ne chevauche pas la réserve
                    _placed = False
                    for rr in range(max_r):
                        for cc in range(max_c):
                            for rows, cols in [(_blocker["rows"], _blocker["cols"]),
                                               (_blocker["cols"], _blocker["rows"])]:
                                if not can_place(rr, cc, rows, cols, x_grid, occ_p, max_r, max_c):
                                    continue
                                _new_cells = {(rr+dr, cc+dc)
                                              for dr in range(rows) for dc in range(cols)}
                                if _new_cells & _reserve:
                                    continue  # chevauchement avec bloc réservé
                                _blocker["r"] = rr; _blocker["c"] = cc
                                _blocker["rows"] = rows; _blocker["cols"] = cols
                                result_p.append(_blocker)
                                occ_p = build_occ(result_p)
                                _placed = True; break
                            if _placed: break
                        if _placed: break
                    if not _placed:
                        # Impossible de déplacer ce blocker -> échec
                        result_p.append(_blocker)  # remettre en place
                        return False, -1, -1

                return True, r, c

            _push_n_ok = _push_n_fail = 0
            _still_sorted = sorted(_still_missing_push, key=lambda b: -(b["rows"]*b["cols"]))
            for _missing_b in _still_sorted:
                _success = False
                for _tr, _tc in [(_missing_b["rows"], _missing_b["cols"]),
                                  (_missing_b["cols"], _missing_b["rows"])]:
                    _ok, _r, _c = _find_best_block_and_push(result_push, _tr, _tc)
                    if _ok:
                        # Placer le manquant dans le bloc libéré
                        occ_p = build_occ(result_push)
                        if can_place(_r, _c, _tr, _tc, x_grid, occ_p, max_r, max_c):
                            result_push.append({**_missing_b, "r": _r, "c": _c,
                                                "rows": _tr, "cols": _tc})
                            _push_n_ok += 1; _success = True; break
                if not _success:
                    _push_n_fail += 1

            if _push_n_fail < best_n_failed:
                best_placed = result_push
                best_n_placed = len(result_push) - len(placed)
                best_n_failed = _push_n_fail

        # ── Fallback smart_place_one si des bâtiments restent manquants ──
    # Le placement en séries peut laisser quelques petits producteurs (2x2)
    # sans position à cause de la fragmentation du terrain.
    # smart_place_one (ancien algo, ordre mixte taille desc) garantit 0 manquants.
    if best_n_failed > 0:
        # Fallback garanti 140/140 : placement en raster pur selon l'ordre
        # Neutres (bords) → petits producteurs <=9 cases → reste taille desc.
        # Le raster pur (sans heuristiques de fragmentation) garantit que
        # les petits producteurs (2x2, 3x3) trouvent leurs blocs contigus.
        _fb_neutres = sorted(
            [b for b in to_place_base if b["type"] == "Neutre"],
            key=lambda b: -(b["rows"] * b["cols"])
        )
        _fb_petits = sorted(
            [b for b in to_place_base
             if b["type"] == "Producteur" and b["rows"] * b["cols"] <= 9],
            key=lambda b: b["rows"] * b["cols"]
        )
        _fb_petits_ids = {id(b) for b in _fb_petits}
        _fb_reste = sorted(
            [b for b in to_place_base
             if b["type"] != "Neutre" and id(b) not in _fb_petits_ids],
            key=lambda b: -(b["rows"] * b["cols"])
        )

        # Raster pur pour les Neutres (bords d'abord)
        _fb_result = [dict(b) for b in placed]
        _fb_ok = _fb_fail = 0
        for _b in _fb_neutres:
            _occ = build_occ(_fb_result); _ok = False
            for _r, _c in cells_border_first:
                for _rows, _cols in [(_b["rows"], _b["cols"]),
                                     (_b["cols"], _b["rows"])]:
                    if can_place(_r, _c, _rows, _cols, x_grid, _occ,
                                 max_r, max_c):
                        _fb_result.append({**_b, "r": _r, "c": _c,
                                           "rows": _rows, "cols": _cols})
                        _fb_ok += 1; _ok = True; break
                if _ok: break
            if not _ok: _fb_fail += 1

        # Raster pur pour petits prods puis reste
        for _b in _fb_petits + _fb_reste:
            _occ = build_occ(_fb_result); _ok = False
            for _rows, _cols in [(_b["rows"], _b["cols"]),
                                  (_b["cols"], _b["rows"])]:
                if _ok: break
                for _r in range(max_r):
                    if _ok: break
                    for _c in range(max_c):
                        if can_place(_r, _c, _rows, _cols, x_grid, _occ,
                                     max_r, max_c):
                            _fb_result.append({**_b, "r": _r, "c": _c,
                                               "rows": _rows, "cols": _cols})
                            _fb_ok += 1; _ok = True; break
            if not _ok: _fb_fail += 1

        def _count_mandatory_fails(result_list):
            """Compte les manquants OBLIGATOIRES dans result_list."""
            _placed_noms = Counter(b["nom"].strip() for b in result_list)
            _n_fail = 0
            for _bd in buildings_def:
                _nom = _bd["nom"].strip()
                _is_mandatory = force_complete or _bd.get("placement", "Obligatoire").strip().lower() != "optionnel"
                if _is_mandatory:
                    _n_fail += max(0, _bd["nombre"] - _placed_noms.get(_nom, 0))
            return _n_fail

        _fb_score  = quick_score(_fb_result)
        _fb_mandatory_fails = _count_mandatory_fails(_fb_result)
        _best_score = quick_score(best_placed) if best_placed else 0
        _best_mandatory_fails = _count_mandatory_fails(best_placed) if best_placed else best_n_failed

        # Règle principale: tout manquant OBLIGATOIRE est éliminatoire.
        # Si le cluster a des manquants obligatoires et que le fallback n'en a pas,
        # on choisit toujours le fallback — quelle que soit la différence de score.
        # Si les deux ont des manquants (terrain vraiment trop plein), on compare
        # les scores pénalisés avec une pénalité très élevée.
        _PENALTY = max(50000, _best_score + _fb_score + 1)  # toujours éliminatoire
        _score_with_penalty_fb   = _fb_score   - _fb_mandatory_fails   * _PENALTY
        _score_with_penalty_best = _best_score - _best_mandatory_fails  * _PENALTY
        if _score_with_penalty_fb > _score_with_penalty_best:
            best_placed   = _fb_result
            best_n_placed = _fb_ok
            best_n_failed = _fb_fail

    # Les bâtiments OPTIONNELS (en mode score maximal) sont retournés séparément.
    # Ils seront placés APRÈS optimize(), dans les espaces libres restants.
    # Ainsi ils n'interfèrent pas avec le cluster optimal.
    return best_placed, best_n_placed, best_n_failed, to_place_optional




# ══════════════════════════════════════════════════════
# OPTIMISEUR
# ══════════════════════════════════════════════════════

def make_x_grid(terrain_grid, max_r, max_c):
    """
    Retourne une grille booléenne où True = case INVALIDE pour le placement.
    Une case est invalide si :
      - elle contient un 'X' (bord du terrain), OU
      - elle est None ET se trouve à l'EXTÉRIEUR du périmètre X
        (terrain non rectangulaire : zone hors de l'enceinte).
    L'extérieur est identifié par flood-fill depuis les bords de la grille.
    """
    from collections import deque

    # Flood-fill depuis tous les bords None pour identifier l'extérieur
    exterior = set()
    queue = deque()

    def _seed(r, c):
        if 0 <= r < max_r and 0 <= c < max_c and terrain_grid[r][c] != "X" and (r, c) not in exterior:
            exterior.add((r, c))
            queue.append((r, c))

    for r in range(max_r):
        _seed(r, 0)
        _seed(r, max_c - 1)
    for c in range(max_c):
        _seed(0, c)
        _seed(max_r - 1, c)

    while queue:
        r, c = queue.popleft()
        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            _seed(r + dr, c + dc)

    # Une case est invalide si X ou extérieure
    return [
        [terrain_grid[r][c] == "X" or (r, c) in exterior
         for c in range(max_c)]
        for r in range(max_r)
    ]


def make_occ_grid(placed, max_r, max_c, exclude_ids=None):
    occ = [[False] * max_c for _ in range(max_r)]
    excl = exclude_ids or set()
    for b in placed:
        if id(b) in excl:
            continue
        for r, c in cells_of(b):
            if 0 <= r < max_r and 0 <= c < max_c:
                occ[r][c] = True
    return occ


def can_place(r, c, rows, cols, x_grid, occ, max_r, max_c):
    if r < 0 or c < 0 or r + rows > max_r or c + cols > max_c:
        return False
    return not any(
        x_grid[r + dr][c + dc] or occ[r + dr][c + dc]
        for dr in range(rows) for dc in range(cols)
    )


def _score_of(b, culturels):
    """Contribution d'un seul batiment producteur au score global."""
    if b["type"] != "Producteur" or b["priorite"] <= 0:
        return 0.0
    cult = culture_received(b, culturels)
    return boost_level(cult, b) * b["priorite"]


def _score_of_cached(b, cult_total_map):
    """
    Contribution d'un producteur au score, en utilisant un dict pre-calcule
    cult_total_map : {id(prod): culture_recue} pour eviter de rappeler culture_received.
    """
    if b["type"] != "Producteur" or b["priorite"] <= 0:
        return 0.0
    cult = cult_total_map.get(id(b), 0.0)
    return boost_level(cult, b) * b["priorite"]


def _build_cult_map(producteurs, culturels):
    """
    Pre-calcule la culture recue par chaque producteur depuis les culturels donnes.
    Retourne {id(prod): culture_recue}.
    """
    # Pre-calculer les zones de rayonnement de chaque culturel (cache)
    zones = {id(c): radiation_zone(c) for c in culturels}
    result = {}
    for p in producteurs:
        pc = cells_of(p)
        total = sum(c["culture"] for c in culturels if pc & zones[id(c)])
        result[id(p)] = total
    return result


def _best_position_for(b, placed, x_grid, max_r, max_c):
    """
    Cherche la meilleure position pour b via un score incremental rapide.
    Toutes les zones de rayonnement sont pre-calculees une seule fois.
    Les Neutres sont ignores immediatement.

    Pour les Culturels: un bonus est appliqué pour couvrir les producteurs
    haute priorité qui n'ont pas encore atteint leur seuil maximum.
    Cela garantit que plus la priorité d'un bâtiment est élevée,
    plus les culturels cherchent à le couvrir.
    """
    if b["type"] == "Neutre":
        return score_placement(placed), None

    orig = (b["r"], b["c"], b["rows"], b["cols"])
    occ  = make_occ_grid(placed, max_r, max_c, exclude_ids={id(b)})
    culturels   = [p for p in placed if p["type"] == "Culturel"]
    producteurs = [p for p in placed if p["type"] == "Producteur"]

    # Culture actuelle de chaque producteur (pre-calculee avec cache de zones)
    cult_map = _build_cult_map(producteurs, culturels)
    orig_score = sum(_score_of_cached(p, cult_map) for p in producteurs)
    best_s, best_pos = orig_score, None

    if b["type"] == "Producteur":
        # Seule la contribution de b change quand on le deplace.
        # Les culturels ne bougent pas -> leurs zones restent identiques.
        # Pour la nouvelle position de b, recalculer sa culture recue.
        contrib_orig = _score_of_cached(b, cult_map)
        base = orig_score - contrib_orig
        # Zones des culturels pre-calculees
        cult_zones = {id(c): radiation_zone(c) for c in culturels}

        # Index des Neutres de MÊME TAILLE que b (pour swaps potentiels)
        # On ne considère que les Neutres de même taille exacte et uniquement
        # pour les Producteurs haute priorité (pour limiter le coût de calcul).
        neutres_meme_taille = []
        if b.get("priorite", 0) > 0:
            for _n in placed:
                if (_n["type"] == "Neutre" and
                        ((_n["rows"] == b["rows"] and _n["cols"] == b["cols"]) or
                         (_n["rows"] == b["cols"] and _n["cols"] == b["rows"]))):
                    neutres_meme_taille.append(_n)

        # OCC sans b (pré-calculé pour les swaps)
        occ_sans_b = make_occ_grid(placed, max_r, max_c, exclude_ids={id(b)})

        for rows, cols in {(b["rows"], b["cols"]), (b["cols"], b["rows"])}:
            for r in range(max_r):
                for c in range(max_c):
                    if (r, c, rows, cols) == orig:
                        continue
                    if can_place(r, c, rows, cols, x_grid, occ, max_r, max_c):
                        # Case libre: déplacement direct
                        b["r"], b["c"], b["rows"], b["cols"] = r, c, rows, cols
                        pc_new = cells_of(b)
                        cult_new = sum(c2["culture"] for c2 in culturels
                                       if pc_new & cult_zones[id(c2)])
                        contrib_new = boost_level(cult_new, b) * b["priorite"] if b["priorite"] > 0 else 0
                        s = base + contrib_new
                        if s > best_s:
                            best_s, best_pos = s, (r, c, rows, cols)
                        b["r"], b["c"], b["rows"], b["cols"] = orig

        # Swaps avec Neutres de même taille (positions actuellement occupées)
        for neutre in neutres_meme_taille:
            nr, nc = neutre["r"], neutre["c"]
            nrows, ncols = neutre["rows"], neutre["cols"]
            # Peut-on placer b à la position du Neutre?
            for rows, cols in {(b["rows"], b["cols"]), (b["cols"], b["rows"])}:
                if rows != nrows or cols != ncols:
                    continue  # taille incompatible pour ce Neutre
                # Vérifier que la position du Neutre est valide pour b
                if any(x_grid[nr+dr][nc+dc] for dr in range(rows) for dc in range(cols)):
                    continue
                occ_sans_both = make_occ_grid(placed, max_r, max_c, exclude_ids={id(b), id(neutre)})
                if not can_place(nr, nc, rows, cols, x_grid, occ_sans_both, max_r, max_c):
                    continue
                # Vérifier que le Neutre peut aller à la position de b
                if not can_place(orig[0], orig[1], nrows, ncols, x_grid, occ_sans_both, max_r, max_c):
                    continue
                # Calculer le gain
                b["r"], b["c"], b["rows"], b["cols"] = nr, nc, rows, cols
                pc_new = cells_of(b)
                cult_new = sum(c2["culture"] for c2 in culturels
                               if pc_new & cult_zones[id(c2)])
                contrib_new = boost_level(cult_new, b) * b["priorite"] if b["priorite"] > 0 else 0
                s = base + contrib_new
                if s > best_s:
                    best_s, best_pos = s, (nr, nc, rows, cols, "swap_neutre", neutre)
                b["r"], b["c"], b["rows"], b["cols"] = orig

    else:  # Culturel
        # b est un culturel. Son deplacement change la culture recue
        # uniquement par les producteurs dans sa zone actuelle ou nouvelle.
        culturels_autres = [p for p in culturels if p is not b]
        # Zones des autres culturels (fixes)
        autres_zones = {id(c): radiation_zone(c) for c in culturels_autres}

        # Cases de chaque producteur (fixes)
        prod_cells = {id(p): cells_of(p) for p in producteurs}

        # Culture de chaque producteur via les AUTRES culturels seulement
        cult_sans_b = {}
        for p in producteurs:
            pc = prod_cells[id(p)]
            cult_sans_b[id(p)] = sum(c["culture"] for c in culturels_autres
                                     if pc & autres_zones[id(c)])

        # Zone actuelle de b et producteurs qu'il couvre
        zone_orig = radiation_zone(b)
        affected_orig = [p for p in producteurs if prod_cells[id(p)] & zone_orig]
        affected_orig_ids = {id(p) for p in affected_orig}

        # Score de base = score sans b (producteurs recalcules avec cult_sans_b)
        base = sum(
            boost_level(cult_sans_b[id(p)], p) * p["priorite"]
            for p in producteurs
        )

        for rows, cols in {(b["rows"], b["cols"]), (b["cols"], b["rows"])}:
            for r in range(max_r):
                for c in range(max_c):
                    if (r, c, rows, cols) == orig:
                        continue
                    if not can_place(r, c, rows, cols, x_grid, occ, max_r, max_c):
                        continue

                    b["r"], b["c"], b["rows"], b["cols"] = r, c, rows, cols
                    zone_new = radiation_zone(b)

                    # Producteurs touches par la nouvelle zone (anciens + nouveaux)
                    affected_new_ids = {id(p) for p in producteurs
                                        if prod_cells[id(p)] & zone_new}
                    all_affected_ids = affected_orig_ids | affected_new_ids
                    all_affected = [p for p in producteurs if id(p) in all_affected_ids]

                    # Pour ces producteurs, recalculer leur score avec b en nouvelle pos
                    delta = 0.0
                    b_cells = cells_of(b)  # deja mis a jour
                    # La zone de b en nouvelle pos = zone_new
                    for p in all_affected:
                        pc = prod_cells[id(p)]
                        # Culture = autres culturels + b si dans zone_new
                        extra = b["culture"] if pc & zone_new else 0.0
                        cult_new_p = cult_sans_b[id(p)] + extra
                        score_new_p = boost_level(cult_new_p, p) * p["priorite"]
                        # Score actuel sans b
                        score_old_p = boost_level(cult_sans_b[id(p)], p) * p["priorite"]
                        delta += score_new_p - score_old_p

                    # Bonus de couverture prioritaire : encourage les culturels
                    # à se placer près des producteurs haute priorité sous-alimentés.
                    # Sans ce bonus, l'algo peut ignorer un producteur prio élevée
                    # si son boost ne change pas immédiatement.
                    _max_prio_local = max((p.get("priorite",0) for p in producteurs), default=1) or 1
                    _coverage_bonus = 0.0
                    for p in producteurs:
                        if p.get("priorite", 0) <= 0 or p.get("boost100", 0) <= 0:
                            continue
                        pc = prod_cells[id(p)]
                        if pc & zone_new:
                            # Culture actuelle de p (avec b en nouvelle pos)
                            _cult_with = cult_sans_b[id(p)] + b["culture"]
                            # Ratio de remplissage vers 100%: plus c'est bas, plus le bonus est fort
                            _ratio_fill = min(1.0, _cult_with / p["boost100"])
                            _prio_weight = p["priorite"] / _max_prio_local
                            # Bonus = culture apportée * prio_normalisée * (1 - ratio_fill)
                            # => grand bonus si bâtiment haute prio et sous-alimenté
                            _coverage_bonus += (b["culture"] * _prio_weight *
                                                (1.0 - _ratio_fill) * 0.01)

                    s = base + delta + _coverage_bonus
                    if s > best_s:
                        best_s, best_pos = s, (r, c, rows, cols)
                    b["r"], b["c"], b["rows"], b["cols"] = orig

    return best_s, best_pos


def _apply_move(b, pos, all_moves):
    if len(pos) >= 5 and pos[4] == "swap_neutre":
        # Swap avec un Neutre: b -> pos[:4], Neutre -> ancienne pos de b
        neutre = pos[5]
        old_b_r, old_b_c = b["r"], b["c"]
        old_b_rows, old_b_cols = b["rows"], b["cols"]
        # Déplacer le Neutre à l'ancienne position de b
        all_moves.append({
            "nom": neutre["nom"],
            "old_r": neutre["r"], "old_c": neutre["c"],
            "old_rows": neutre["rows"], "old_cols": neutre["cols"],
            "new_r": old_b_r, "new_c": old_b_c,
            "new_rows": old_b_rows, "new_cols": old_b_cols,
        })
        neutre["r"], neutre["c"] = old_b_r, old_b_c
        neutre["rows"], neutre["cols"] = old_b_rows, old_b_cols
        # Déplacer b
        all_moves.append({
            "nom": b["nom"],
            "old_r": old_b_r, "old_c": old_b_c,
            "old_rows": old_b_rows, "old_cols": old_b_cols,
            "new_r": pos[0], "new_c": pos[1],
            "new_rows": pos[2], "new_cols": pos[3],
        })
        b["r"], b["c"], b["rows"], b["cols"] = pos[0], pos[1], pos[2], pos[3]
    else:
        all_moves.append({
            "nom":      b["nom"],
            "old_r":    b["r"],    "old_c":    b["c"],
            "old_rows": b["rows"], "old_cols": b["cols"],
            "new_r":    pos[0],    "new_c":    pos[1],
            "new_rows": pos[2],    "new_cols": pos[3],
        })
        b["r"], b["c"], b["rows"], b["cols"] = pos[0], pos[1], pos[2], pos[3]


def _culture_coverage(culturel, placed):
    """Nombre de producteurs couverts par ce culturel."""
    zone = radiation_zone(culturel)
    return sum(1 for b in placed if b["type"] == "Producteur" and cells_of(b) & zone)


def _dist(b1, b2):
    """Distance Manhattan entre centres de deux batiments."""
    r1 = b1["r"] + b1["rows"] / 2
    c1 = b1["c"] + b1["cols"] / 2
    r2 = b2["r"] + b2["rows"] / 2
    c2 = b2["c"] + b2["cols"] / 2
    return abs(r1 - r2) + abs(c1 - c2)


def optimize(placed, terrain_grid, max_r, max_c, n_passes=2, progress_cb=None, time_budget_sec=None):
    """
    Optimisation en 3 etapes :

    Etape 1 (greedy convergence) : chaque batiment cherche sa meilleure position
      globale en score direct. Ordre de la liste = culturels puis producteurs,
      ce qui permet aux culturels de s'installer et aux producteurs de se grouper.
      Repete n_passes fois jusqu'a convergence.

    Etape 2 (rapprochement agressif des culturels inutiles) : les culturels qui
      ne couvrent aucun producteur sont deplaces AU PLUS PRES des producteurs
      les moins bien alimentes, meme si le gain de score est nul.
      L'idee : rompre l'optimum local pour permettre a l'etape 3 d'aller plus loin.

    Etape 3 (reconvergence finale) : une nouvelle passe greedy complete apres
      les deplacements forcees de l'etape 2.

    Repete les etapes 2+3 jusqu'a ce qu'il n'y ait plus de culturels inutiles
    ou qu'aucun progres ne soit possible.
    """
    x_grid = make_x_grid(terrain_grid, max_r, max_c)
    placed = [dict(b) for b in placed]
    all_moves = []

    n = len(placed)
    total_ops = n_passes * n * 2 + 5 * n
    op = [0]

    def tick(k=1):
        op[0] += k
        if progress_cb:
            progress_cb(min(op[0] / max(total_ops, 1), 0.98))

    def greedy_pass(max_inner=10):
        """
        Passe greedy : chaque batiment Culturel ou Producteur cherche sa meilleure
        position. Les Neutres sont ignores (ils ne contribuent pas au score).
        max_inner limite le nombre d'iterations internes pour eviter les boucles longues.
        """
        for _ in range(max_inner):
            improved = False
            # Trier par priorité décroissante : les casernes (prio=10) bougent en premier.
            # Ainsi les culturels se déplacent ensuite autour des casernes bien placées.
            sorted_placed = sorted(
                [b for b in placed if b["type"] != "Neutre"],
                key=lambda b: -b.get("priorite", 0)
            )
            for b in sorted_placed:
                best_s, best_pos = _best_position_for(b, placed, x_grid, max_r, max_c)
                if best_pos:
                    _apply_move(b, best_pos, all_moves)
                    improved = True
                tick()
            if not improved:
                break

    # ── Etape 1 ──
    for _ in range(n_passes):
        greedy_pass()

    # ── Etapes 2+3 : boucle de deblocage ──
    for _outer in range(n_passes + 1):
        culturels = [b for b in placed if b["type"] == "Culturel"]
        producteurs = [b for b in placed if b["type"] == "Producteur"]
        inutiles = [c for c in culturels if _culture_coverage(c, placed) == 0]

        if not inutiles:
            break

        # Pour chaque culturel inutile, trouver la position la plus proche
        # d'un producteur peu couvert qui soit libre, et l'y deplacer.
        prod_by_cult = sorted(producteurs, key=lambda p: culture_received(p, culturels))
        any_forced = False

        for cult in sorted(inutiles, key=lambda c: c["culture"], reverse=True):
            orig_cult = (cult["r"], cult["c"], cult["rows"], cult["cols"])
            occ = make_occ_grid(placed, max_r, max_c, exclude_ids={id(cult)})

            # Chercher d'abord s'il existe une position ameliorant le score
            # (sans contrainte de proximite)
            best_improve_s, best_improve_pos = score_placement(placed), None
            for rows, cols in {(cult["rows"], cult["cols"]), (cult["cols"], cult["rows"])}:
                for r in range(max_r):
                    for c in range(max_c):
                        if (r, c, rows, cols) == orig_cult: continue
                        if not can_place(r, c, rows, cols, x_grid, occ, max_r, max_c): continue
                        cult["r"], cult["c"], cult["rows"], cult["cols"] = r, c, rows, cols
                        s = score_placement(placed)
                        if s > best_improve_s:
                            best_improve_s, best_improve_pos = s, (r, c, rows, cols)
                        cult["r"], cult["c"], cult["rows"], cult["cols"] = orig_cult

            if best_improve_pos:
                # Il existe une position qui ameliore le score : on l'applique
                _apply_move(cult, best_improve_pos, all_moves)
                any_forced = True
                continue

            # Sinon : deplacement force vers le producteur le moins bien alimente
            # On choisit la position libre la plus proche de ce producteur
            best_dist = float("inf")
            best_forced_pos = None
            target_prod = prod_by_cult[0]  # producteur avec le moins de culture

            for rows, cols in {(cult["rows"], cult["cols"]), (cult["cols"], cult["rows"])}:
                for r in range(max_r):
                    for c in range(max_c):
                        if (r, c, rows, cols) == orig_cult: continue
                        if not can_place(r, c, rows, cols, x_grid, occ, max_r, max_c): continue
                        # Distance entre le culturel (si place en r,c) et le producteur cible
                        cult["r"], cult["c"], cult["rows"], cult["cols"] = r, c, rows, cols
                        # Verifier que le rayonnement couvrirait le producteur
                        zone = radiation_zone(cult)
                        prod_cells = cells_of(target_prod)
                        if zone & prod_cells:
                            # Position qui couvre directement -> priorite absolue
                            d = -1
                        else:
                            d = _dist(cult, target_prod)
                        if d < best_dist:
                            best_dist, best_forced_pos = d, (r, c, rows, cols)
                        cult["r"], cult["c"], cult["rows"], cult["cols"] = orig_cult

            if best_forced_pos and best_forced_pos != orig_cult:
                _apply_move(cult, best_forced_pos, all_moves)
                any_forced = True
            tick()

        if not any_forced:
            break

        # ── Etape 3 : reconvergence apres deplacements forces ──
        greedy_pass()

    # ── Etape finale : swaps de paires de Culturels ──
    # L'optimiseur greedy peut etre bloque dans un optimum local ou
    # aucun deplacement individuel n'ameliore le score.
    # Echanger simultanement deux Culturels peut debloquer la situation :
    # chacun occupe la position de l'autre, ce qui peut mieux couvrir
    # les producteurs dans les deux zones.
    culturels_list = [b for b in placed if b["type"] == "Culturel"]
    producteurs_list = [b for b in placed if b["type"] == "Producteur"]
    prod_cells_cache = {id(p): cells_of(p) for p in producteurs_list}

    swap_improved = True
    n_swap_passes = 0
    max_swap_passes = n_passes  # Plus de passes = meilleur résultat
    while swap_improved and n_swap_passes < max_swap_passes:
        swap_improved = False
        for i, c1 in enumerate(culturels_list):
            for c2 in culturels_list[i+1:]:
                zone1 = radiation_zone(c1)
                zone2 = radiation_zone(c2)
                # Producteurs dont la couverture changera
                affected = [p for p in producteurs_list
                            if prod_cells_cache[id(p)] & zone1
                            or prod_cells_cache[id(p)] & zone2]
                if not affected:
                    continue

                score_before = sum(
                    boost_level(culture_received(p, culturels_list), p) * p["priorite"]
                    for p in affected
                )

                # Tester le swap : c1 prend la position de c2 et vice versa.
                # Chaque bâtiment CONSERVE ses propres dimensions (rows, cols) :
                # les dimensions sont des propriétés fixes du bâtiment dans le jeu.
                # On teste aussi les pivots individuels si le bâtiment n'est pas carré.
                old1 = (c1["r"], c1["c"], c1["rows"], c1["cols"])
                old2 = (c2["r"], c2["c"], c2["rows"], c2["cols"])
                best_delta = 0.0
                best_config = None

                # Générer les variantes: chaque bâtiment peut rester dans son orientation
                # ou pivoter (seulement si rows != cols, i.e. pas carré)
                c1_orientations = [(c1["rows"], c1["cols"])]
                if c1["rows"] != c1["cols"]:
                    c1_orientations.append((c1["cols"], c1["rows"]))
                c2_orientations = [(c2["rows"], c2["cols"])]
                if c2["rows"] != c2["cols"]:
                    c2_orientations.append((c2["cols"], c2["rows"]))

                occ_without = make_occ_grid(placed, max_r, max_c,
                                            exclude_ids={id(c1), id(c2)})

                for r1_new, co1_new in c1_orientations:
                    for r2_new, co2_new in c2_orientations:
                        # c1 va à la position de c2 (avec son orientation r1_new x co1_new)
                        # c2 va à la position de c1 (avec son orientation r2_new x co2_new)
                        new_r1, new_c1 = old2[0], old2[1]  # position de c2
                        new_r2, new_c2 = old1[0], old1[1]  # position de c1

                        # Vérifier que les dimensions tiennent dans les cases disponibles
                        if not can_place(new_r1, new_c1, r1_new, co1_new,
                                         x_grid, occ_without, max_r, max_c):
                            continue
                        if not can_place(new_r2, new_c2, r2_new, co2_new,
                                         x_grid, occ_without, max_r, max_c):
                            continue

                        # Vérifier que les deux nouvelles positions ne se chevauchent
                        # pas entre elles (occ_without les exclut toutes les deux,
                        # donc can_place ne détecte pas leur conflit mutuel)
                        _cells1 = {(new_r1+_dr, new_c1+_dc)
                                   for _dr in range(r1_new) for _dc in range(co1_new)}
                        _cells2 = {(new_r2+_dr, new_c2+_dc)
                                   for _dr in range(r2_new) for _dc in range(co2_new)}
                        if _cells1 & _cells2:
                            continue

                        new1 = (new_r1, new_c1, r1_new, co1_new)
                        new2 = (new_r2, new_c2, r2_new, co2_new)

                        c1["r"], c1["c"], c1["rows"], c1["cols"] = new1
                        c2["r"], c2["c"], c2["rows"], c2["cols"] = new2

                        score_after = sum(
                            boost_level(culture_received(p, culturels_list), p) * p["priorite"]
                            for p in affected
                        )
                        delta = score_after - score_before
                        if delta > best_delta:
                            best_delta = delta
                            best_config = (new1, new2)

                        c1["r"], c1["c"], c1["rows"], c1["cols"] = old1
                        c2["r"], c2["c"], c2["rows"], c2["cols"] = old2

                if best_config:
                    c1["r"], c1["c"], c1["rows"], c1["cols"] = best_config[0]
                    c2["r"], c2["c"], c2["rows"], c2["cols"] = best_config[1]
                    all_moves.append({
                        "nom": c1["nom"], "old_r": old1[0], "old_c": old1[1],
                        "old_rows": old1[2], "old_cols": old1[3],
                        "new_r": best_config[0][0], "new_c": best_config[0][1],
                        "new_rows": best_config[0][2], "new_cols": best_config[0][3],
                    })
                    all_moves.append({
                        "nom": c2["nom"], "old_r": old2[0], "old_c": old2[1],
                        "old_rows": old2[2], "old_cols": old2[3],
                        "new_r": best_config[1][0], "new_c": best_config[1][1],
                        "new_rows": best_config[1][2], "new_cols": best_config[1][3],
                    })
                    swap_improved = True
        # Apres une passe de swaps, relancer la greedy pour consolider
        if swap_improved:
            greedy_pass()
        n_swap_passes += 1

    # ── Etape bonus : boost ciblé des producteurs haute priorité sous-boostés ──
    # Cherche les producteurs haute prio qui n'ont pas atteint 100% et tente
    # de déplacer des culturels pour les couvrir davantage.
    _max_prio = max((b["priorite"] for b in placed if b["type"] == "Producteur"), default=1)
    _prio_thresh = _max_prio * 0.5

    for _boost_pass in range(n_passes):
        _culturels = [b for b in placed if b["type"] == "Culturel"]
        _producteurs = [b for b in placed if b["type"] == "Producteur"]
        # Producteurs haute prio n'ayant pas atteint 100%
        _targets = sorted(
            [p for p in _producteurs
             if p["priorite"] >= _prio_thresh
             and boost_level(culture_received(p, _culturels), p) < 100],
            key=lambda p: -(p["priorite"] * (p["boost100"] - culture_received(p, _culturels)))
        )
        if not _targets:
            break
        _improved_bonus = False
        for _target in _targets:
            _score_before = score_placement(placed)
            # Trier les culturels par couverture actuelle (inutiles en premier)
            _cults_sorted = sorted(
                _culturels,
                key=lambda c: (
                    1 if cells_of(_target) & radiation_zone(c) else 0,
                    culture_received(_target, _culturels)
                )
            )
            for _cult in _cults_sorted:
                _orig = (_cult["r"], _cult["c"], _cult["rows"], _cult["cols"])
                _occ = make_occ_grid(placed, max_r, max_c, exclude_ids={id(_cult)})
                _best_s, _best_pos = _score_before, None
                # Chercher une position qui couvre le target ET améliore le score
                _target_cells = cells_of(_target)
                for _rows, _cols in {(_cult["rows"], _cult["cols"]), (_cult["cols"], _cult["rows"])}:
                    for _r in range(max_r):
                        for _c in range(max_c):
                            if (_r, _c, _rows, _cols) == _orig:
                                continue
                            if not can_place(_r, _c, _rows, _cols, x_grid, _occ, max_r, max_c):
                                continue
                            _cult["r"], _cult["c"], _cult["rows"], _cult["cols"] = _r, _c, _rows, _cols
                            _zone_new = radiation_zone(_cult)
                            if _target_cells & _zone_new:
                                _s = score_placement(placed)
                                if _s > _best_s:
                                    _best_s, _best_pos = _s, (_r, _c, _rows, _cols)
                            _cult["r"], _cult["c"], _cult["rows"], _cult["cols"] = _orig
                if _best_pos:
                    _apply_move(_cult, _best_pos, all_moves)
                    _improved_bonus = True
                    break
        if _improved_bonus:
            greedy_pass()
        else:
            break

    # ── Simulated Annealing : débloquer les optima locaux ──
    # S'active quand le score des casernes est < 70% du maximum théorique.
    # Budget temps = n_passes * 25s (ex: 3 passes -> 75s, 8 passes -> 200s).
    import math as _math, random as _random, time as _time_sa_mod

    _culturels_sa  = [b for b in placed if b["type"] == "Culturel"]
    _prods_sa      = [b for b in placed if b["type"] == "Producteur"]
    _max_prio_sa   = max((b["priorite"] for b in _prods_sa), default=1)
    _prio_thresh_sa = _max_prio_sa * 0.8
    _top_prods_sa  = [b for b in _prods_sa if b.get("priorite", 0) >= _prio_thresh_sa]

    _top_score_now_sa = sum(
        boost_level(culture_received(p, _culturels_sa), p) * p["priorite"]
        for p in _top_prods_sa
    )
    _top_score_max_sa = sum(100 * p["priorite"] for p in _top_prods_sa)

    if _top_prods_sa and _top_score_now_sa < 0.70 * _top_score_max_sa and n_passes >= 2:
        # Budget SA: soit défini par time_budget_sec (temps total - marge greedy),
        # soit estimé à partir du nombre de passes.
        if time_budget_sec is not None:
            _sa_budget = max(10, time_budget_sec - n_passes * 12)  # réserver ~12s/passe pour greedy
        else:
            _sa_budget = n_passes * 25        # 25s de SA par passe
        _sa_seed     = 42
        _random.seed(_sa_seed)
        # Refroidissement calibré: T=800 -> T=1 sur toute la durée
        # À ~200 iters/s, iters_total ≈ 200 * _sa_budget
        _iters_est   = 200 * _sa_budget
        _T_sa        = 800.0
        _T_min_sa    = 0.5
        _alpha_sa    = (_T_min_sa / _T_sa) ** (1.0 / max(_iters_est, 1))
        _t0_sa       = _time_sa_mod.time()

        _sa_current  = [dict(b) for b in placed]
        _sa_score    = score_placement(_sa_current)
        _sa_best     = [dict(b) for b in _sa_current]
        _sa_best_score = _sa_score
        _sa_n_iter   = _sa_n_improve = _sa_last_improve = 0
        _REHEAT_N    = max(8000, _iters_est // 6)

        # Inclure TOUS les bâtiments dans les swaps (y compris les Neutres).
        # Un Neutre peut occuper une position stratégique qu'un Producteur
        # haute priorité devrait avoir -> swap bénéfique.
        _non_neut_idx = list(range(len(_sa_current)))
        _top_idx      = [i for i, b in enumerate(_sa_current)
                         if b["type"] == "Producteur" and b.get("priorite", 0) >= _prio_thresh_sa]

        def _sa_occ(pl, excl):
            _o = [[False]*max_c for _ in range(max_r)]
            for _b in pl:
                if id(_b) in excl: continue
                for _dr in range(_b["rows"]):
                    for _dc in range(_b["cols"]):
                        _rr,_cc=_b["r"]+_dr,_b["c"]+_dc
                        if 0<=_rr<max_r and 0<=_cc<max_c: _o[_rr][_cc]=True
            return _o

        while _T_sa > _T_min_sa and _time_sa_mod.time()-_t0_sa < _sa_budget:
            # Réchauffe adaptative si bloqué
            if _sa_n_iter - _sa_last_improve > _REHEAT_N:
                _T_sa = max(_T_sa * 5.0, 50.0)
                _sa_last_improve = _sa_n_iter

            # 15% big move: déplacer une caserne vers sa meilleure position locale
            if _random.random() < 0.15 and _top_idx:
                _ip  = _random.choice(_top_idx)
                _bp  = _sa_current[_ip]
                _sv  = (_bp["r"],_bp["c"],_bp["rows"],_bp["cols"])
                _ow  = _sa_occ(_sa_current, {id(_bp)})
                _cls = [b for b in _sa_current if b["type"]=="Culturel"]
                _czn = {id(c): radiation_zone(c) for c in _cls}
                _co  = boost_level(sum(c["culture"] for c in _cls
                                       if cells_of(_bp) & _czn[id(c)]), _bp) * _bp.get("priorite",0)
                _bpp = None; _bc = _co
                for _rw,_cw in [(_bp["rows"],_bp["cols"]),(_bp["cols"],_bp["rows"])]:
                    for _r in range(max_r):
                        for _c in range(max_c):
                            if (_r,_c,_rw,_cw)==_sv: continue
                            if not can_place(_r,_c,_rw,_cw,x_grid,_ow,max_r,max_c): continue
                            _bp["r"],_bp["c"],_bp["rows"],_bp["cols"]=_r,_c,_rw,_cw
                            _pcc=cells_of(_bp)
                            _cv=sum(_cv2["culture"] for _cv2 in _cls if _pcc&_czn[id(_cv2)])
                            _ct=boost_level(_cv,_bp)*_bp.get("priorite",0)
                            if _ct>_bc: _bc=_ct; _bpp=(_r,_c,_rw,_cw)
                            _bp["r"],_bp["c"],_bp["rows"],_bp["cols"]=_sv
                if _bpp:
                    _bp["r"],_bp["c"],_bp["rows"],_bp["cols"]=_bpp
                    _ns=score_placement(_sa_current)
                    if _ns>=_sa_score:
                        _sa_score=_ns
                        if _ns>_sa_best_score:
                            _sa_best_score=_ns; _sa_best=[dict(b) for b in _sa_current]
                            _sa_n_improve+=1; _sa_last_improve=_sa_n_iter
                    else:
                        _bp["r"],_bp["c"],_bp["rows"],_bp["cols"]=_sv
                _T_sa*=_alpha_sa; _sa_n_iter+=1; continue

            # 85% swap de deux bâtiments non-neutres
            if len(_non_neut_idx)<2: _T_sa*=_alpha_sa; _sa_n_iter+=1; continue
            _i1,_i2=_random.sample(_non_neut_idx,2)
            _b1,_b2=_sa_current[_i1],_sa_current[_i2]
            _s1=(_b1["r"],_b1["c"],_b1["rows"],_b1["cols"])
            _s2=(_b2["r"],_b2["c"],_b2["rows"],_b2["cols"])
            _ow2=_sa_occ(_sa_current,{id(_b1),id(_b2)})
            _sw=False; _np1=_np2=None
            for _r1,_co1 in [(_b1["rows"],_b1["cols"]),(_b1["cols"],_b1["rows"])]:
                if _sw: break
                if not can_place(_s2[0],_s2[1],_r1,_co1,x_grid,_ow2,max_r,max_c): continue
                _ow3=[_row[:] for _row in _ow2]
                for _dr in range(_r1):
                    for _dc in range(_co1):
                        _rr,_cc=_s2[0]+_dr,_s2[1]+_dc
                        if 0<=_rr<max_r and 0<=_cc<max_c: _ow3[_rr][_cc]=True
                for _r2,_co2 in [(_b2["rows"],_b2["cols"]),(_b2["cols"],_b2["rows"])]:
                    if not can_place(_s1[0],_s1[1],_r2,_co2,x_grid,_ow3,max_r,max_c): continue
                    _c1s={(_s2[0]+_dr,_s2[1]+_dc) for _dr in range(_r1) for _dc in range(_co1)}
                    _c2s={(_s1[0]+_dr,_s1[1]+_dc) for _dr in range(_r2) for _dc in range(_co2)}
                    if _c1s&_c2s: continue
                    _np1=(_s2[0],_s2[1],_r1,_co1); _np2=(_s1[0],_s1[1],_r2,_co2)
                    _sw=True; break
                if _sw: break
            if not _sw: _T_sa*=_alpha_sa; _sa_n_iter+=1; continue
            _b1["r"],_b1["c"],_b1["rows"],_b1["cols"]=_np1
            _b2["r"],_b2["c"],_b2["rows"],_b2["cols"]=_np2
            _ns=score_placement(_sa_current); _dl=_ns-_sa_score
            if _dl>0 or _random.random()<_math.exp(min(_dl/_T_sa,0)):
                _sa_score=_ns
                if _ns>_sa_best_score:
                    _sa_best_score=_ns; _sa_best=[dict(b) for b in _sa_current]
                    _sa_n_improve+=1; _sa_last_improve=_sa_n_iter
            else:
                _b1["r"],_b1["c"],_b1["rows"],_b1["cols"]=_s1
                _b2["r"],_b2["c"],_b2["rows"],_b2["cols"]=_s2
            _T_sa*=_alpha_sa; _sa_n_iter+=1

        # Restaurer le meilleur état SA dans placed
        _best_map={(_b["nom"],_b.get("_id",j)): _b for j,_b in enumerate(_sa_best)}
        for _j, _b in enumerate(placed):
            if _j < len(_sa_best):
                _b.update({k:v for k,v in _sa_best[_j].items() if k in ("r","c","rows","cols")})

        # Reconverger après SA
        greedy_pass()

        # ── Phase de perturbation : forcer le cluster si on part d'un raster ──
    # Si l'optimum local est pauvre (peu de casernes à 100%),
    # tenter de forcer les producteurs haute prio vers le centre
    # puis reconverger. Brise le deadlock raster -> cluster.
    _culturels_now = [b for b in placed if b["type"] == "Culturel"]
    _prods_now = [b for b in placed if b["type"] == "Producteur"]
    _max_prio_now = max((b["priorite"] for b in _prods_now), default=1)
    _prio_thresh_now = _max_prio_now * 0.8
    _top_prods_now = [b for b in _prods_now if b.get("priorite", 0) >= _prio_thresh_now]

    # Calculer le score courant des casernes haute prio
    _top_score_now = sum(
        boost_level(culture_received(p, _culturels_now), p) * p["priorite"]
        for p in _top_prods_now
    )
    # Score max possible pour ces casernes si toutes à 100%
    _top_score_max = sum(100 * p["priorite"] for p in _top_prods_now)
    # Si on n'a atteint que < 60% du max, tenter la perturbation
    _ratio = _top_score_now / max(_top_score_max, 1)

    if _ratio < 0.60 and n_passes >= 2:
        _tr, _tc = max_r / 2.0, max_c / 2.0

        for _perturb_round in range(n_passes):
            # Étape A: forcer chaque caserne haute prio vers le centre
            _top_sorted = sorted(_top_prods_now,
                                 key=lambda b: -b.get("boost100", 0))
            for _prod in _top_sorted:
                _occ = make_occ_grid(placed, max_r, max_c, exclude_ids={id(_prod)})
                _best_d = float("inf")
                _best_center = None
                for _r in range(max_r):
                    for _c in range(max_c):
                        for _rows, _cols in [(_prod["rows"], _prod["cols"]),
                                              (_prod["cols"], _prod["rows"])]:
                            if can_place(_r, _c, _rows, _cols, x_grid, _occ, max_r, max_c):
                                _d = abs(_r + _rows/2 - _tr) + abs(_c + _cols/2 - _tc)
                                if _d < _best_d:
                                    _best_d = _d
                                    _best_center = (_r, _c, _rows, _cols)
                if _best_center and (_best_center[0] != _prod["r"] or _best_center[1] != _prod["c"]):
                    _apply_move(_prod, _best_center, all_moves)

            # Étape B: forcer les culturels inutiles vers les casernes
            _culturels_now2 = [b for b in placed if b["type"] == "Culturel"]
            _inutiles = [c for c in _culturels_now2
                         if not any(cells_of(p) & radiation_zone(c) for p in _top_prods_now)]
            _inutiles_sorted = sorted(_inutiles,
                                      key=lambda c: -(c["culture"] * max(c.get("rayonnement", 1), 1)))

            for _cult in _inutiles_sorted:
                _occ = make_occ_grid(placed, max_r, max_c, exclude_ids={id(_cult)})
                # Chercher la position la plus proche d'une caserne peu couverte
                _least_covered = min(
                    _top_prods_now,
                    key=lambda p: culture_received(p, [b for b in placed if b["type"]=="Culturel"])
                )
                _target_cr = _least_covered["r"] + _least_covered["rows"] / 2.0
                _target_cc = _least_covered["c"] + _least_covered["cols"] / 2.0
                _best_d2 = float("inf")
                _best_pos2 = None
                for _r in range(max_r):
                    for _c in range(max_c):
                        for _rows, _cols in [(_cult["rows"], _cult["cols"]),
                                              (_cult["cols"], _cult["rows"])]:
                            if not can_place(_r, _c, _rows, _cols, x_grid, _occ, max_r, max_c):
                                continue
                            # Vérifier que la zone couvre la caserne cible
                            _r0, _c0 = _r, _c
                            _r1, _c1 = _r+_rows-1, _c+_cols-1
                            _ray = _cult.get("rayonnement", 1)
                            _zone = {(rz, cz)
                                     for rz in range(_r0-_ray, _r1+_ray+1)
                                     for cz in range(_c0-_ray, _c1+_ray+1)
                                     if not (_r0<=rz<=_r1 and _c0<=cz<=_c1)
                                     and 0<=rz<max_r and 0<=cz<max_c}
                            if cells_of(_least_covered) & _zone:
                                _d2 = abs(_r+_rows/2-_target_cr)+abs(_c+_cols/2-_target_cc)
                                if _d2 < _best_d2:
                                    _best_d2 = _d2
                                    _best_pos2 = (_r, _c, _rows, _cols)
                if _best_pos2 and (_best_pos2[0] != _cult["r"] or _best_pos2[1] != _cult["c"]):
                    _apply_move(_cult, _best_pos2, all_moves)

            # Étape C: reconverger
            greedy_pass()

            # Vérifier si on a progressé
            _culturels_now = [b for b in placed if b["type"] == "Culturel"]
            _new_top_score = sum(
                boost_level(culture_received(p, _culturels_now), p) * p["priorite"]
                for p in _top_prods_now
            )
            if _new_top_score / max(_top_score_max, 1) >= 0.60:
                break  # Objectif atteint

    return placed, all_moves


# ══════════════════════════════════════════════════════
# GENERATION DU FICHIER EXCEL DE SORTIE
# ══════════════════════════════════════════════════════

def build_excel_output(optimized, original_placed, terrain_grid, max_r, max_c, buildings_def):
    from collections import Counter
    wb = openpyxl.Workbook()
    culturels = [b for b in optimized if b["type"] == "Culturel"]
    orig_culturels = [b for b in original_placed if b["type"] == "Culturel"]

    # ─────────────────────────────────────
    # ONGLET 1 : Liste batiments
    # ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Liste batiments"
    headers = ["Nom", "Type", "Placé", "Coord (L,C)", "Orientation",
               "Priorité", "Placement", "Culture reçue", "Boost atteint", "Score boost"]
    widths  = [32, 12, 8, 12, 12, 10, 12, 14, 13, 12]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        style_header(ws1.cell(1, ci), h)
        ws1.column_dimensions[get_column_letter(ci)].width = w

    row_i = 2
    for b in sorted(optimized, key=lambda x: (x["type"], x["nom"])):
        cult = culture_received(b, culturels) if b["type"] == "Producteur" else 0
        boost = boost_level(cult, b)
        prio = b["priorite"]
        score = boost * prio if b["type"] == "Producteur" else ""
        qte_boost = b["quantite"] * (1 + boost / 100) if b["type"] == "Producteur" else ""
        orient = "H" if b["cols"] >= b["rows"] else "V"
        fill = mfill(C_ORANGE if b["type"] == "Culturel" else C_GREEN if b["type"] == "Producteur" else C_GRAY)
        prio_disp = int(prio) if prio == int(prio) else prio
        placement_val = b.get("placement", "Obligatoire")
        vals = [b["nom"], b["type"],
                "Oui",
                f"{get_column_letter(b['c']+1)}{b['r']+1}", orient,
                prio_disp if b["type"] == "Producteur" else "",
                placement_val,
                round(cult, 1), f"{boost}%",
                round(score, 3) if score != "" else ""]
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row_i, ci, v)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_i += 1

    # Ajouter les bâtiments non placés (optionnels ET obligatoires) avec Placé=Non
    placed_noms = Counter(b["nom"].strip() for b in optimized)
    for b_def in buildings_def:
        nom = b_def["nom"].strip()
        placed_count = placed_noms.get(nom, 0)
        missing = b_def["nombre"] - placed_count
        if missing > 0:
            is_optional = b_def.get("placement", "Obligatoire").strip().lower() == "optionnel"
            for _ in range(missing):
                placement_val = b_def.get("placement", "Obligatoire")
                # Fond rouge pâle pour les obligatoires manquants, gris pour les optionnels
                fill_np = mfill("FFFFD7D7") if not is_optional else mfill(C_GRAY)
                vals_np = [nom, b_def["type"],
                           "Non", "", "",
                           "", placement_val,
                           "", "", ""]
                for ci, v in enumerate(vals_np, 1):
                    cell = ws1.cell(row_i, ci, v)
                    cell.fill = fill_np
                    cell.border = thin_border()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if ci == 3:  # colonne Placé
                        cell.font = Font(bold=True, color="FFCC0000")
                row_i += 1

    # ─────────────────────────────────────
    # ONGLET 2 : Synthese
    # ─────────────────────────────────────
    ws2 = wb.create_sheet("Synthese")

    # ── Section 1 : Score global et boosts par type de batiment ──
    score_avant = score_placement(original_placed)
    score_apres = score_placement(optimized)

    # En-tete section 1
    titre1 = ws2.cell(1, 1, "Boosts par type de batiment producteur")
    titre1.font = Font(bold=True, size=12)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    titre1.alignment = Alignment(horizontal="center")
    titre1.fill = mfill("FF1F4E79")
    titre1.font = Font(bold=True, size=12, color=C_WHITE)

    hdrs_boost = ["Type de batiment", "Priorite",
                  "Avant : 0%", "Avant : 25%", "Avant : 50%", "Avant : 100%",
                  "Apres : 0%", "Apres : 25%", "Apres : 50%", "Apres : 100%"]
    widths_boost = [36, 10, 11, 11, 11, 12, 11, 11, 11, 12]
    for ci, (h, w) in enumerate(zip(hdrs_boost, widths_boost), 1):
        style_header(ws2.cell(2, ci), h)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    # Colonnes score global
    ws2.column_dimensions["B"].width = max(ws2.column_dimensions["B"].width, 18)
    ws2.column_dimensions["C"].width = max(ws2.column_dimensions["C"].width, 18)
    ws2.column_dimensions["D"].width = max(ws2.column_dimensions["D"].width, 18)

    # Calculer boosts avant/apres par nom de batiment unique
    def boost_counts(placed_list, cult_list):
        """Retourne dict {nom: {0:n, 25:n, 50:n, 100:n, priorite:p}}"""
        counts = {}
        for b in placed_list:
            if b["type"] != "Producteur":
                continue
            nom = b["nom"]
            cult = culture_received(b, cult_list)
            boost = boost_level(cult, b)
            prio = b.get("priorite", 0)
            if nom not in counts:
                counts[nom] = {0: 0, 25: 0, 50: 0, 100: 0, "priorite": prio}
            else:
                # Garder la vraie priorité (pas 0 par défaut)
                if prio > counts[nom]["priorite"]:
                    counts[nom]["priorite"] = prio
            counts[nom][boost] += 1
        return counts

    counts_avant = boost_counts(original_placed, orig_culturels)
    counts_apres = boost_counts(optimized, culturels)

    # Union de tous les noms de batiments producteurs
    all_prod_names = sorted(set(list(counts_avant.keys()) + list(counts_apres.keys())))

    C_BOOST0   = "FFFFD7D7"  # rouge pale  = pas de boost
    C_BOOST25  = "FFFFF2CC"  # jaune pale  = 25%
    C_BOOST50  = "FFD9EAD3"  # vert pale   = 50%
    C_BOOST100 = "FF93C47D"  # vert vif    = 100%
    boost_colors = {0: C_BOOST0, 25: C_BOOST25, 50: C_BOOST50, 100: C_BOOST100}

    row_i = 3
    for nom in all_prod_names:
        av = counts_avant.get(nom, {0: 0, 25: 0, 50: 0, 100: 0, "priorite": 0})
        ap = counts_apres.get(nom, {0: 0, 25: 0, 50: 0, 100: 0, "priorite": 0})
        prio = av.get("priorite") or ap.get("priorite") or 0

        vals = [nom, prio,
                av[0], av[25], av[50], av[100],
                ap[0], ap[25], ap[50], ap[100]]

        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row_i, ci, v)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Colorier les colonnes de boost
            if ci == 3:  cell.fill = mfill(C_BOOST0)
            elif ci == 4: cell.fill = mfill(C_BOOST25)
            elif ci == 5: cell.fill = mfill(C_BOOST50)
            elif ci == 6: cell.fill = mfill(C_BOOST100)
            elif ci == 7: cell.fill = mfill(C_BOOST0)
            elif ci == 8: cell.fill = mfill(C_BOOST25)
            elif ci == 9: cell.fill = mfill(C_BOOST50)
            elif ci == 10: cell.fill = mfill(C_BOOST100)
            # Mettre en gras les valeurs ameliorees
            if ci in (7, 8, 9, 10):
                boost_val = [0, 25, 50, 100][ci - 7]
                avant_val = av[boost_val]
                apres_val = ap[boost_val]
                if boost_val > 0 and apres_val > avant_val:
                    cell.font = Font(bold=True, color="FF006400")
                elif boost_val > 0 and apres_val < avant_val:
                    cell.font = Font(bold=True, color="FFCC0000")
        row_i += 1

    # Ligne de total
    total_row = row_i
    ws2.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    ws2.cell(total_row, 1).fill = mfill("FFD9D9D9")
    ws2.cell(total_row, 1).border = thin_border()
    ws2.cell(total_row, 2).border = thin_border()
    for ci, col_boost in enumerate([0, 25, 50, 100], 3):
        # Avant
        total_av = sum(counts_avant.get(n, {col_boost: 0})[col_boost] for n in all_prod_names)
        cell = ws2.cell(total_row, ci, total_av)
        cell.font = Font(bold=True)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center")
        cell.fill = mfill(boost_colors[col_boost])
        # Apres
        total_ap = sum(counts_apres.get(n, {col_boost: 0})[col_boost] for n in all_prod_names)
        cell2 = ws2.cell(total_row, ci + 4, total_ap)
        cell2.font = Font(bold=True)
        cell2.border = thin_border()
        cell2.alignment = Alignment(horizontal="center")
        cell2.fill = mfill(boost_colors[col_boost])
    row_i = total_row + 2

    # ── Section 2 : Score global ──
    titre2 = ws2.cell(row_i, 1, "Score de boost de production")
    titre2.font = Font(bold=True, size=12, color=C_WHITE)
    titre2.fill = mfill("FF1F4E79")
    titre2.alignment = Alignment(horizontal="center")
    ws2.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
    row_i += 1

    for ci, h in enumerate(["", "Score avant", "Score apres", "Gain"], 1):
        cell = ws2.cell(row_i, ci, h)
        if h:
            style_header(cell, h)
        ws2.column_dimensions[get_column_letter(ci)].width = max(
            ws2.column_dimensions[get_column_letter(ci)].width, 18)
    row_i += 1

    delta_score = score_apres - score_avant
    ws2.cell(row_i, 1, "Score boost global").font = Font(bold=True)
    ws2.cell(row_i, 1).border = thin_border()
    ws2.cell(row_i, 1).fill = mfill("FFD9D9D9")

    cell_av = ws2.cell(row_i, 2, round(score_avant, 2))
    cell_av.font = Font(bold=True)
    cell_av.border = thin_border()
    cell_av.alignment = Alignment(horizontal="center")

    cell_ap = ws2.cell(row_i, 3, round(score_apres, 2))
    cell_ap.font = Font(bold=True)
    cell_ap.border = thin_border()
    cell_ap.alignment = Alignment(horizontal="center")

    cell_gain = ws2.cell(row_i, 4, round(delta_score, 2))
    cell_gain.font = Font(bold=True,
                          color=C_GAIN if delta_score >= 0 else C_LOSS)
    cell_gain.border = thin_border()
    cell_gain.alignment = Alignment(horizontal="center")
    row_i += 2


    # ─────────────────────────────────────
    # ONGLET 3 : Deplacements
    # ─────────────────────────────────────
    ws3 = wb.create_sheet("Deplacements")
    hdrs3 = ["#", "Batiment", "Position initiale", "Position finale", "Sequence d'operations"]
    widths3 = [4, 30, 16, 16, 70]
    for ci, (h, w) in enumerate(zip(hdrs3, widths3), 1):
        style_header(ws3.cell(1, ci), h)
        ws3.column_dimensions[get_column_letter(ci)].width = w

    orig_pos = {}
    for b in original_placed:
        orig_pos.setdefault(b["nom"], []).append((b["r"], b["c"], b["rows"], b["cols"]))

    real_moves = []
    used_orig = {nom: 0 for nom in orig_pos}
    for b in optimized:
        nom = b["nom"]
        if nom not in orig_pos:
            continue
        idx = used_orig.get(nom, 0)
        if idx < len(orig_pos[nom]):
            op = orig_pos[nom][idx]
            used_orig[nom] = idx + 1
            if op[0] != b["r"] or op[1] != b["c"]:
                real_moves.append({
                    "nom": nom,
                    "old_r": op[0], "old_c": op[1],
                    "old_rows": op[2], "old_cols": op[3],
                    "new_r": b["r"], "new_c": b["c"],
                    "new_rows": b["rows"], "new_cols": b["cols"],
                })

    if not real_moves:
        ws3.cell(2, 1, "Aucun deplacement effectue - placement deja optimal.")
    else:
        # ── Séquençage intelligent avec mise en réserve ──
        # Chaque bâtiment reçoit un _uid unique pour distinguer les instances
        # de bâtiments de même nom (ex: 12x "Site culturel compact").
        for _i, _b in enumerate(original_placed):
            _b["_uid"] = _i
        for _mv in real_moves:
            # Retrouver le _uid de l'instance initiale correspondante
            _match = next((b for b in original_placed
                           if b["nom"] == _mv["nom"]
                           and b["r"] == _mv["old_r"] and b["c"] == _mv["old_c"]), None)
            _mv["_uid"] = _match["_uid"] if _match else -1

        # Grille des positions initiales (cellule -> bâtiment par _uid)
        _init_cells = {}  # (r,c) -> bâtiment initial
        for b in original_placed:
            for dr in range(b["rows"]):
                for dc in range(b["cols"]):
                    _init_cells[(b["r"]+dr, b["c"]+dc)] = b

        # État courant
        _freed_cells = set()       # cellules libérées par un déplacement déjà effectué
        _in_reserve = {}           # _uid -> mv pour les bâtiments en réserve
        _done_uids = set()         # _uid des bâtiments déjà replacés

        def _target_cells(mv):
            return {(mv["new_r"]+dr, mv["new_c"]+dc)
                    for dr in range(mv["new_rows"]) for dc in range(mv["new_cols"])}

        def _source_cells(mv):
            return {(mv["old_r"]+dr, mv["old_c"]+dc)
                    for dr in range(mv["old_rows"]) for dc in range(mv["old_cols"])}

        def _blockers(mv):
            """Instances initiales qui occupent la cible et n'ont pas encore été déplacées."""
            tc = _target_cells(mv)
            seen = {}
            for cell in tc - _freed_cells:
                if cell in _init_cells:
                    b = _init_cells[cell]
                    # Blocker = toute instance DIFFÉRENTE de celle qu'on déplace
                    if b["_uid"] != mv["_uid"] and b["_uid"] not in seen:
                        seen[b["_uid"]] = b
            return list(seen.values())

        # Générer la séquence d'opérations
        operations = []
        pending = list(real_moves)
        iterations = 0

        while pending and iterations < len(real_moves) * 6:
            iterations += 1
            progress = False
            remaining = []

            for mv in pending:
                if mv["_uid"] in _done_uids:
                    progress = True
                    continue

                blockers = _blockers(mv)

                if not blockers:
                    # Cible libre: déplacement possible
                    from_str = (f"{get_column_letter(mv['old_c']+1)}{mv['old_r']+1}"
                                if mv["_uid"] not in _in_reserve else "Réserve")
                    new_str = f"{get_column_letter(mv['new_c']+1)}{mv['new_r']+1}"
                    note = "(retour de réserve)" if mv["_uid"] in _in_reserve else ""
                    operations.append({
                        "type": "move",
                        "nom": mv["nom"],
                        "from": from_str,
                        "to": new_str,
                        "note": note,
                    })
                    _freed_cells.update(_source_cells(mv))
                    _freed_cells -= _target_cells(mv)
                    _done_uids.add(mv["_uid"])
                    _in_reserve.pop(mv["_uid"], None)
                    progress = True
                else:
                    # Cible bloquée: mettre chaque bloqueur en réserve
                    for blocker in blockers:
                        buid = blocker["_uid"]
                        if buid not in _in_reserve and buid not in _done_uids:
                            old_b = f"{get_column_letter(blocker['c']+1)}{blocker['r']+1}"
                            # Label distinctif si plusieurs instances de même nom
                            same_name_count = sum(1 for b in original_placed
                                                   if b["nom"] == blocker["nom"])
                            bat_label = (f"{blocker['nom']} (en {old_b})"
                                         if same_name_count > 1 else blocker["nom"])
                            operations.append({
                                "type": "reserve",
                                "nom": blocker["nom"],
                                "bat_label": bat_label,
                                "from": old_b,
                                "to": "Réserve",
                                "note": f"(libère la place pour '{mv['nom']}')",
                            })
                            # Libérer ses cellules
                            for dr in range(blocker["rows"]):
                                for dc in range(blocker["cols"]):
                                    _freed_cells.add((blocker["r"]+dr, blocker["c"]+dc))
                            _in_reserve[buid] = blocker
                        progress = True
                    remaining.append(mv)

            pending = remaining
            if not progress:
                break

        # Bâtiments encore en réserve non replacés (ils ont été mis en réserve
        # mais leur déplacement final n'était pas dans real_moves)
        for buid, blocker in list(_in_reserve.items()):
            if buid not in _done_uids:
                # Retrouver le déplacement correspondant
                mv_final = next((mv for mv in real_moves if mv["_uid"] == buid), None)
                if mv_final:
                    new_str = f"{get_column_letter(mv_final['new_c']+1)}{mv_final['new_r']+1}"
                else:
                    # Pas de déplacement prévu: replacer à la position initiale
                    new_str = f"{get_column_letter(blocker['c']+1)}{blocker['r']+1}"
                same_name_count = sum(1 for b in original_placed
                                       if b["nom"] == blocker["nom"])
                from_str = (f"{get_column_letter(blocker['c']+1)}{blocker['r']+1}"
                            if same_name_count > 1 else "")
                operations.append({
                    "type": "move",
                    "nom": blocker["nom"],
                    "bat_label": (f"{blocker['nom']} (ex {from_str})"
                                  if same_name_count > 1 else blocker["nom"]),
                    "from": "Réserve",
                    "to": new_str,
                    "note": "(retour de réserve)",
                })

        # ── Écriture dans l'onglet ──
        # Couleurs selon le type d'opération
        C_RESERVE = "FFFFF2CC"   # jaune pâle: mise en réserve
        C_RETURN  = "FFD9EAD3"   # vert pâle: retour de réserve
        C_MOVE    = "FFCFE2F3"   # bleu pâle: déplacement simple

        for si, op in enumerate(operations):
            ri = si + 2
            color = (C_RESERVE if op["type"] == "reserve" else
                     C_RETURN  if "réserve" in op["note"] else C_MOVE)
            _label = op.get("bat_label", op["nom"])
            if op["type"] == "reserve":
                action = (f"Mettre '{_label}' en réserve (retirer de {op['from']}) "
                          f"{op['note']}")
            else:
                action = (f"Placer '{_label}' en {op['to']}"
                          f"{' depuis ' + op['from'] if op['from'] != 'Réserve' else ' (depuis la réserve)'}"
                          f". {op['note']}")
            ws3.cell(ri, 1, si + 1)
            ws3.cell(ri, 2, op["nom"])
            ws3.cell(ri, 3, op["from"])
            ws3.cell(ri, 4, op["to"])
            ws3.cell(ri, 5, action)
            ws3.cell(ri, 5).alignment = Alignment(wrap_text=True, vertical="top")
            ws3.row_dimensions[ri].height = 40
            for ci in range(1, 6):
                cell = ws3.cell(ri, ci)
                cell.border = thin_border()
                cell.fill = mfill(color)
                if ci < 5:
                    cell.alignment = Alignment(horizontal="center", vertical="top")

        # Légende en bas
        leg_row = len(operations) + 3
        ws3.cell(leg_row, 1, "Légende :").font = Font(bold=True)
        for i, (label, color) in enumerate([
                ("Déplacement direct", C_MOVE),
                ("Mise en réserve temporaire", C_RESERVE),
                ("Retour de réserve", C_RETURN)], 1):
            cell = ws3.cell(leg_row + i, 1, label)
            cell.fill = mfill(color)
            cell.border = thin_border()

    # ─────────────────────────────────────
    # ONGLET 4 : Terrain optimise (carte)
    # ─────────────────────────────────────
    ws4 = wb.create_sheet("Terrain optimise")

    placed_grid = {}
    for b in optimized:
        for dr in range(b["rows"]):
            for dc in range(b["cols"]):
                placed_grid[(b["r"]+dr, b["c"]+dc)] = b

    # Identifier les bâtiments déplacés (position finale ≠ position initiale)
    orig_pos_map = {}
    for b in original_placed:
        orig_pos_map.setdefault(b["nom"], []).append((b["r"], b["c"]))
    _used_orig = {nom: 0 for nom in orig_pos_map}
    moved_bat_ids = set()
    for b in optimized:
        nom = b["nom"]
        if nom in orig_pos_map:
            idx_o = _used_orig.get(nom, 0)
            if idx_o < len(orig_pos_map[nom]):
                orig_r, orig_c = orig_pos_map[nom][idx_o]
                _used_orig[nom] = idx_o + 1
                if orig_r != b["r"] or orig_c != b["c"]:
                    moved_bat_ids.add(id(b))
        else:
            # Bâtiment qui n'existait pas dans l'original = nouvellement placé
            moved_bat_ids.add(id(b))

    col_w = 14
    row_h = 20
    for r in range(max_r):
        ws4.row_dimensions[r+1].height = row_h
    for c in range(max_c):
        ws4.column_dimensions[get_column_letter(c+1)].width = col_w

    for r in range(max_r):
        for c in range(max_c):
            cell = ws4.cell(r+1, c+1)
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                continue
            if terrain_grid[r][c] == "X":
                cell.value = "X"
                cell.fill = mfill(C_BORDX)
                cell.font = Font(bold=True, color=C_WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif (r, c) in placed_grid:
                b = placed_grid[(r, c)]
                fill_hex = (C_ORANGE if b["type"] == "Culturel"
                            else C_GREEN if b["type"] == "Producteur"
                            else C_GRAY)
                cell.fill = mfill(fill_hex)
                cell.border = thin_border()
                if b["r"] == r and b["c"] == c:
                    cult = culture_received(b, culturels) if b["type"] == "Producteur" else 0
                    boost = boost_level(cult, b)
                    label = b["nom"]
                    if b["type"] == "Producteur" and boost > 0:
                        label += f"\n+{boost}%"
                    cell.value = label
                    cell.alignment = Alignment(horizontal="center", vertical="center",
                                               wrap_text=True)
                    # Nom en rouge si le bâtiment a été déplacé, sinon couleur normale
                    is_moved = id(b) in moved_bat_ids
                    cell.font = Font(bold=(boost > 0),
                                     color="FFCC0000" if is_moved else "FF000000")
                    if b["rows"] > 1 or b["cols"] > 1:
                        try:
                            ws4.merge_cells(
                                start_row=r+1, start_column=c+1,
                                end_row=r+b["rows"], end_column=c+b["cols"]
                            )
                        except Exception:
                            pass

    # Legende
    leg_r = max_r + 2
    ws4.cell(leg_r, 1, "Legende").font = Font(bold=True)
    for i, (label, color) in enumerate([
            ("Culturel", C_ORANGE), ("Producteur", C_GREEN), ("Neutre", C_GRAY)], 1):
        cell = ws4.cell(leg_r+i, 1, label)
        cell.fill = mfill(color)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center")
    # Note sur les noms en rouge
    ws4.cell(leg_r+4, 1, "Nom en rouge = bâtiment déplacé").font = Font(italic=True, color="FFCC0000")

    # ─────────────────────────────────────
    # ONGLET 5 : Terrain initial
    # ─────────────────────────────────────
    ws5 = wb.create_sheet("Terrain initial")

    # Construire la grille des bâtiments initialement placés
    placed_grid_init = {}
    for b in original_placed:
        for dr in range(b["rows"]):
            for dc in range(b["cols"]):
                placed_grid_init[(b["r"]+dr, b["c"]+dc)] = b

    for r in range(max_r):
        ws5.row_dimensions[r+1].height = row_h
    for c in range(max_c):
        ws5.column_dimensions[get_column_letter(c+1)].width = col_w

    orig_culturels_init = [b for b in original_placed if b["type"] == "Culturel"]
    for r in range(max_r):
        for c in range(max_c):
            cell = ws5.cell(r+1, c+1)
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                continue
            if terrain_grid[r][c] == "X":
                cell.value = "X"
                cell.fill = mfill(C_BORDX)
                cell.font = Font(bold=True, color=C_WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif (r, c) in placed_grid_init:
                b = placed_grid_init[(r, c)]
                fill_hex = (C_ORANGE if b["type"] == "Culturel"
                            else C_GREEN if b["type"] == "Producteur"
                            else C_GRAY)
                cell.fill = mfill(fill_hex)
                cell.border = thin_border()
                if b["r"] == r and b["c"] == c:
                    cult = culture_received(b, orig_culturels_init) if b["type"] == "Producteur" else 0
                    boost = boost_level(cult, b)
                    label = b["nom"]
                    if b["type"] == "Producteur" and boost > 0:
                        label += f"\n+{boost}%"
                    cell.value = label
                    cell.alignment = Alignment(horizontal="center", vertical="center",
                                               wrap_text=True)
                    cell.font = Font(bold=(boost > 0))
                    if b["rows"] > 1 or b["cols"] > 1:
                        try:
                            ws5.merge_cells(
                                start_row=r+1, start_column=c+1,
                                end_row=r+b["rows"], end_column=c+b["cols"]
                            )
                        except Exception:
                            pass

    # Légende
    leg_r5 = max_r + 2
    ws5.cell(leg_r5, 1, "Legende").font = Font(bold=True)
    for i, (label, color) in enumerate([
            ("Culturel", C_ORANGE), ("Producteur", C_GREEN), ("Neutre", C_GRAY)], 1):
        cell = ws5.cell(leg_r5+i, 1, label)
        cell.fill = mfill(color)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()

    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════
# INTERFACE STREAMLIT
# ══════════════════════════════════════════════════════

# Initialisation du session_state : les resultats persistent entre les reruns
if "result_excel" not in st.session_state:
    st.session_state.result_excel        = None
    st.session_state.score_init          = None
    st.session_state.score_opt           = None
    st.session_state.moved_summary       = []
    st.session_state.last_filename       = None


# ── Choix de la source de données ──
_input_mode = st.radio(
    "Source du fichier",
    ["📁 Fichier Excel (.xlsx)", "🔗 Google Sheets (URL publique)"],
    horizontal=True
)

uploaded = None
_sheets_url = None

if _input_mode == "📁 Fichier Excel (.xlsx)":
    uploaded = st.file_uploader(
        "Choisissez votre fichier Excel de ville (.xlsx)",
        type=["xlsx"],
        help="Le fichier doit contenir un onglet Terrain et un onglet Batiments."
    )
else:
    _sheets_url_input = st.text_input(
        "URL Google Sheets",
        placeholder="https://docs.google.com/spreadsheets/d/...",
        help=(
            "Le fichier Google Sheets doit être accessible publiquement "
            "(Partage > Tout le monde avec le lien peut consulter).\n\n"
            "Comment partager : Fichier > Partager > Partager avec d'autres personnes "
            "> Tout le monde avec le lien > Lecteur."
        )
    )
    if _sheets_url_input and "docs.google.com/spreadsheets" in _sheets_url_input:
        try:
            import requests as _requests
            # Extraire l'ID du sheet
            import re as _re
            _match = _re.search("/spreadsheets/d/([a-zA-Z0-9_-]+)", _sheets_url_input)
            if _match:
                _sheet_id = _match.group(1)
                # Extraire gid si présent (onglet spécifique)
                _gid_match = _re.search("[#&]gid=([0-9]+)", _sheets_url_input)
                _export_url = f"https://docs.google.com/spreadsheets/d/{_sheet_id}/export?format=xlsx"
                with st.spinner("Téléchargement depuis Google Sheets..."):
                    _resp = _requests.get(_export_url, timeout=15)
                if _resp.status_code == 200:
                    import io as _io
                    uploaded = _io.BytesIO(_resp.content)
                    uploaded.name = "google_sheets.xlsx"
                    _sheets_url = _sheets_url_input
                    st.success(f"✅ Fichier Google Sheets chargé ({len(_resp.content)//1024} Ko)")
                else:
                    st.error(
                        f"❌ Impossible de télécharger le fichier (code {_resp.status_code}). "
                        "Vérifiez que le fichier est partagé publiquement (lecture seule)."
                    )
            else:
                st.error("URL Google Sheets non reconnue. Copiez l'URL complète depuis la barre d'adresse.")
        except Exception as _e:
            st.error(f"Erreur lors du téléchargement : {_e}")
    elif _sheets_url_input:
        st.warning("Cette URL ne semble pas être un Google Sheets. Utilisez l'URL depuis la barre d'adresse de votre navigateur.")

# Initialiser _sheets_url si pas défini (mode fichier local)
if '_sheets_url' not in dir():
    _sheets_url = None

# Si un nouveau fichier est charge, on efface les resultats precedents
if uploaded:
    if uploaded.name != st.session_state.last_filename:
        st.session_state.result_excel  = None
        st.session_state.score_init    = None
        st.session_state.score_opt     = None
        st.session_state.moved_summary = []
        st.session_state.last_filename = uploaded.name

    try:
        wb_in = openpyxl.load_workbook(uploaded)
    except Exception as e:
        st.error(f"Impossible de lire le fichier : {e}")
        st.stop()

    sheet_names = wb_in.sheetnames
    if len(sheet_names) < 2:
        st.error("Le fichier doit contenir au moins 2 onglets (Terrain + Batiments).")
        st.stop()

    terrain_ws = wb_in[sheet_names[0]]
    bat_ws     = wb_in[sheet_names[1]]

    terrain_grid, max_r, max_c = read_terrain(terrain_ws)
    buildings_def = read_buildings_def(bat_ws)
    placed = enrich(read_placed_buildings(terrain_ws), buildings_def)
    original_placed = [dict(b) for b in placed]

    n_culturels   = sum(1 for b in placed if b["type"] == "Culturel")
    n_producteurs = sum(1 for b in placed if b["type"] == "Producteur")
    n_neutres     = sum(1 for b in placed if b["type"] == "Neutre")
    score_init    = score_placement(placed)

    st.success(f"Fichier charge : **{len(placed)} batiments** sur un terrain **{max_r} x {max_c}**")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Score initial", f"{score_init:.2f}")
    col2.metric("Batiments culturels", n_culturels)
    col3.metric("Batiments producteurs", n_producteurs)
    col4.metric("Batiments neutres", n_neutres)

    with st.expander("Detail de la culture initiale par producteur"):
        culturels_init = [b for b in placed if b["type"] == "Culturel"]
        for b in sorted(placed, key=lambda x: x["nom"]):
            if b["type"] == "Producteur":
                cult = culture_received(b, culturels_init)
                boost = boost_level(cult, b)
                st.write(
                    f"**{b['nom']}** - Culture recue : {cult:.0f} / "
                    f"Seuil 25% : {b['boost25']:.0f} | 50% : {b['boost50']:.0f} | 100% : {b['boost100']:.0f} "
                    f"→ **Boost : {boost}%**"
                )

    st.divider()

    # ── Batiments manquants ──
    placed_counts = {}
    for b in placed:
        placed_counts[b["nom"].strip()] = placed_counts.get(b["nom"].strip(), 0) + 1
    missing_list = []
    missing_optional_list = []
    for b_def in buildings_def:
        nom = b_def["nom"].strip()
        needed = b_def["nombre"] - placed_counts.get(nom, 0)
        if needed > 0:
            is_opt = b_def.get("placement", "Obligatoire").strip().lower() == "optionnel"
            entry = f"**{nom}** : {needed} à placer"
            if is_opt:
                missing_optional_list.append(entry + " *(Optionnel)*")
            else:
                missing_list.append(entry)

    if missing_list or missing_optional_list:
        label = f"⚠️ {len(missing_list)+len(missing_optional_list)} type(s) de bâtiments non encore placés sur le terrain"
        with st.expander(label):
            if missing_list:
                st.markdown("**Obligatoires :**")
                for m in missing_list:
                    st.write(m)
            if missing_optional_list:
                st.markdown("**Optionnels *(peuvent rester non placés en mode Score maximal)* :**")
                for m in missing_optional_list:
                    st.write(m)
    do_place_missing = True  # toujours placer automatiquement

    temps_max_min = 3
    temps_max_sec = temps_max_min * 60
    n_passes = max(1, temps_max_sec // 40)
    st.caption(f"≈ {n_passes} passes d'optimisation")
    force_complete = False  # toujours mode score maximal

    if st.button("Lancer l'optimisation", type="primary"):
        # Effacer les resultats precedents avant de relancer
        st.session_state.result_excel  = None
        st.session_state.score_opt     = None
        st.session_state.moved_summary = []

        progress_bar = st.progress(0)
        status = st.empty()

        # Placer les batiments manquants si demande
        placed_for_optim = placed
        n_placed_new = 0
        n_failed_new = 0
        optionnels_a_placer = []  # bâtiments optionnels à tenter après optimize()
        # Appeler place_missing_buildings si des bâtiments manquent (obligatoires OU optionnels).
        # Les optionnels sont toujours tentés après l'optimisation.
        _has_missing = bool(missing_list or missing_optional_list)
        if do_place_missing and _has_missing:
            status.info("Placement des batiments manquants (plusieurs essais)...")
            n_trials = min(3, n_passes)
            placed_for_optim, n_placed_new, n_failed_new, optionnels_a_placer = place_missing_buildings(
                placed, buildings_def, terrain_grid, max_r, max_c, n_trials=n_trials,
                force_complete=force_complete
            )
        elif not _has_missing:
            # Tous les bâtiments obligatoires sont déjà sur le terrain,
            # mais les optionnels doivent quand même être tentés après optimisation.
            placed_for_optim, _, _, optionnels_a_placer = place_missing_buildings(
                placed, buildings_def, terrain_grid, max_r, max_c, n_trials=1,
                force_complete=False
            )
            if n_failed_new > 0:
                from collections import Counter as _Cnt
                placed_names = _Cnt(b["nom"].strip() for b in placed_for_optim)
                fail_types = _Cnt()
                fail_names_oblig = []
                fail_names_opt   = []
                for b_def in buildings_def:
                    nom = b_def["nom"].strip()
                    short = b_def["nombre"] - placed_names.get(nom, 0)
                    if short > 0:
                        fail_types[b_def["type"]] += short
                        is_opt = b_def.get("placement", "Obligatoire").strip().lower() == "optionnel"
                        entry = f"{nom} ({short}×)"
                        if is_opt:
                            fail_names_opt.append(entry)
                        else:
                            fail_names_oblig.append(entry)

                if fail_names_opt and not fail_names_oblig:
                    st.info(
                        f"ℹ️ **{len(fail_names_opt)} bâtiment(s) Optionnel(s) non placés** "
                        f"(terrain plein — autorisé en mode Score maximal). "
                        f"Bâtiments : {', '.join(fail_names_opt)}."
                    )
                elif fail_types.get("Neutre", 0) == n_failed_new:
                    st.warning(
                        f"⚠️ **{n_failed_new} bâtiment(s) Neutre(s) n'ont pas pu être placés** "
                        f"(terrain trop fragmenté). **Cela n'affecte pas le score.** "
                        f"Bâtiments : {', '.join(fail_names_oblig + fail_names_opt)}."
                    )
                else:
                    msg = f"⚠️ {n_failed_new} bâtiment(s) n'ont pas pu être placés (terrain plein)."
                    if fail_names_oblig:
                        msg += f" **Obligatoires manquants** : {', '.join(fail_names_oblig)}."
                    if fail_names_opt:
                        msg += f" Optionnels non placés (normal en mode Score maximal) : {', '.join(fail_names_opt)}."
                    st.warning(msg)
            if n_placed_new > 0:
                st.success(f"✅ {n_placed_new} bâtiment(s) placés sur le terrain.")

        status.info("Optimisation en cours... Veuillez patienter.")

        def update_prog(v):
            progress_bar.progress(v)

        optimized, moves = optimize(
            placed_for_optim, terrain_grid, max_r, max_c,
            n_passes=n_passes, progress_cb=update_prog,
            time_budget_sec=temps_max_sec
        )

        # Placer les bâtiments OPTIONNELS post-optimisation.
        # Règle: placer le maximum d'optionnels possibles,
        # en cherchant pour chaque optionnel la MEILLEURE position
        # (celle qui maximise le gain de score, ou à défaut couvre
        # le plus de producteurs sous-alimentés).
        # Ne pas placer si cela diminue le score.
        if optionnels_a_placer:
            _xg_opt = make_x_grid(terrain_grid, max_r, max_c)
            _opt_quota = {b["nom"].strip(): b["nombre"] for b in buildings_def
                          if b.get("placement","Obligatoire").strip().lower() == "optionnel"}

            # ── Helpers ──────────────────────────────────────────────────
            def _best_pos_opt(opt_b, current):
                """Meilleure position pour un optionnel: max couverture producteurs,
                puis première case libre. Retourne (pos, gain) ou (None, 0)."""
                _occ = make_occ_grid(current, max_r, max_c)
                _sn  = score_placement(current)
                if opt_b["type"] != "Culturel":
                    for _rw,_cl in [(opt_b["rows"],opt_b["cols"]),(opt_b["cols"],opt_b["rows"])]:
                        for _r in range(max_r):
                            for _c in range(max_c):
                                if can_place(_r,_c,_rw,_cl,_xg_opt,_occ,max_r,max_c):
                                    return (_r,_c,_rw,_cl), 0
                    return None, 0
                _pr = [b for b in current if b["type"] == "Producteur"]
                _cu = [b for b in current if b["type"] == "Culturel"]
                _pc = {id(p): cells_of(p) for p in _pr}
                _bp = None; _bc = -1
                for _rw,_cl in [(opt_b["rows"],opt_b["cols"]),(opt_b["cols"],opt_b["rows"])]:
                    for _r in range(max_r - _rw + 1):
                        for _c in range(max_c - _cl + 1):
                            if not can_place(_r,_c,_rw,_cl,_xg_opt,_occ,max_r,max_c): continue
                            _bat = {**opt_b,"r":_r,"c":_c,"rows":_rw,"cols":_cl}
                            _z   = radiation_zone(_bat)
                            _cov = sum(opt_b["culture"] * max(0, 1 - culture_received(p,_cu)/p["boost100"])
                                       for p in _pr if _pc[id(p)] & _z and p.get("boost100",0) > 0)
                            if _cov > _bc: _bc = _cov; _bp = (_r,_c,_rw,_cl)
                if _bp is None:
                    for _rw,_cl in [(opt_b["rows"],opt_b["cols"]),(opt_b["cols"],opt_b["rows"])]:
                        for _r in range(max_r - _rw + 1):
                            for _c in range(max_c - _cl + 1):
                                if can_place(_r,_c,_rw,_cl,_xg_opt,_occ,max_r,max_c):
                                    _bp = (_r,_c,_rw,_cl); break
                            if _bp: break
                        if _bp: break
                if _bp is None: return None, 0
                _bat = {**opt_b,"r":_bp[0],"c":_bp[1],"rows":_bp[2],"cols":_bp[3]}
                return _bp, score_placement(current + [_bat]) - _sn

            # ── Phase 1 : placement par couverture ────────────────────────
            _opts_sorted = sorted(optionnels_a_placer,
                key=lambda b: (0 if b["type"]=="Culturel" else 1, -(b["rows"]*b["cols"])))
            for _ob in _opts_sorted:
                _pos, _gain = _best_pos_opt(_ob, optimized)
                if _pos is not None and _gain >= 0:
                    optimized.append({**_ob,"r":_pos[0],"c":_pos[1],
                                      "rows":_pos[2],"cols":_pos[3]})

            # ── Phase 2 : compaction ─────────────────────────────────────────────
            # Boucle externe: relancer si des gains ont été faits (nouvelles opportunités).
            # Boucle interne: tester chaque HdV, OCC fraîche, break+restart dès un gain.
            _compact_total_added = 0
            for _outer in range(10):  # max 10 passes externes
                _gains_this_outer = 0
                for _cp in range(200):  # passes internes
                    _placed_c = {}
                    for _b_cnt in optimized:
                        _k = _b_cnt["nom"].strip()
                        _placed_c[_k] = _placed_c.get(_k, 0) + 1
                    _unplaced = [dict(b) for _nom,_tot in _opt_quota.items()
                                 for _ in range(_tot - _placed_c.get(_nom, 0))
                                 for b in optionnels_a_placer if b["nom"].strip() == _nom]
                    if not _unplaced: break

                    _un0 = _unplaced[0]
                    _ur, _uc = _un0["rows"], _un0["cols"]
                    _s_cur = score_placement(optimized)
                    _found_in_pass = False

                    for _h in list(b for b in optimized if b["nom"].strip() in _opt_quota):
                        _wo = [b for b in optimized if b is not _h]
                        _ow = make_occ_grid(_wo, max_r, max_c)
                        _bm = None

                        for _nr in range(max_r - _h["rows"] + 1):
                            if _bm: break
                            for _nc in range(max_c - _h["cols"] + 1):
                                if _nr == _h["r"] and _nc == _h["c"]: continue
                                if not can_place(_nr,_nc,_h["rows"],_h["cols"],_xg_opt,_ow,max_r,max_c): continue
                                _mv = _wo + [{**_h,"r":_nr,"c":_nc}]
                                if score_placement(_mv) < _s_cur - 0.1: continue
                                _om = make_occ_grid(_mv, max_r, max_c)
                                for _rw2,_cl2 in [(_ur,_uc),(_uc,_ur)]:
                                    for _r4 in range(max_r - _rw2 + 1):
                                        for _c4 in range(max_c - _cl2 + 1):
                                            if can_place(_r4,_c4,_rw2,_cl2,_xg_opt,_om,max_r,max_c):
                                                _bm = (_nr,_nc,_r4,_c4,_rw2,_cl2); break
                                        if _bm: break
                                    if _bm: break
                                if _bm: break
                            if _bm: break

                        if _bm:
                            _nr,_nc,_r4,_c4,_rw2,_cl2 = _bm
                            _wo2 = [b for b in optimized if b is not _h]
                            optimized = _wo2 + [{**_h,"r":_nr,"c":_nc}] +                                         [{**_un0,"r":_r4,"c":_c4,"rows":_rw2,"cols":_cl2}]
                            _s_cur = score_placement(optimized)
                            _found_in_pass = True
                            _gains_this_outer += 1
                            _compact_total_added += 1
                            break

                    if not _found_in_pass:
                        break  # Convergé pour cette passe externe

                if _gains_this_outer == 0:
                    break  # Aucun gain même en relançant: vraiment convergé
            # ── Passe greedy finale ───────────────────────────────────────
            if any(b["type"] == "Culturel" for b in optionnels_a_placer):
                _opt_greedy, _ = optimize(optimized, terrain_grid, max_r, max_c, n_passes=1)
                if score_placement(_opt_greedy) >= score_placement(optimized):
                    optimized = _opt_greedy
        progress_bar.progress(1.0)
        # Compter les HdV ajoutés par compaction pour affichage
        status.success("Optimisation terminee !")

        score_opt = score_placement(optimized)

        # Calculer la liste des deplacements reels
        # original_placed = etat avant placement des manquants + avant optimisation
        orig_map = {}
        for b in original_placed:
            orig_map.setdefault(b["nom"], []).append((b["r"], b["c"]))
        used = {n: 0 for n in orig_map}
        summary_lines = []
        for b in optimized:
            nom = b["nom"]
            if nom in orig_map:
                idx = used[nom]
                if idx < len(orig_map[nom]):
                    used[nom] += 1
                    op = orig_map[nom][idx]
                    if op[0] != b["r"] or op[1] != b["c"]:
                        cult_val = culture_received(b, [x for x in optimized if x["type"] == "Culturel"])
                        boost    = boost_level(cult_val, b)
                        icon = "🟠" if b["type"] == "Culturel" else "🟢" if b["type"] == "Producteur" else "⬜"
                        line = (
                            f"{icon} **{nom}** : "
                            f"{get_column_letter(op[1]+1)}{op[0]+1} → {get_column_letter(b['c']+1)}{b['r']+1}"
                            + (f" | Boost apres : **{boost}%**" if b["type"] == "Producteur" else "")
                        )
                        summary_lines.append(line)

        # Generer le fichier Excel et stocker dans session_state
        # Pour l'onglet Deplacements, on compare toujours l'etat du fichier INPUT
        # (original_placed) avec l'etat optimise final.
        # Si le terrain etait (partiellement) vide, les nouveaux batiments places
        # n'ont pas de "position initiale" -> ils n'apparaissent pas dans Deplacements.
        with st.spinner("Generation du fichier Excel..."):
            output_buf = build_excel_output(
                optimized, original_placed, terrain_grid, max_r, max_c, buildings_def
            )
            st.session_state.result_excel  = output_buf.getvalue()
            st.session_state.score_init    = score_placement(placed_for_optim)
            st.session_state.score_opt     = score_opt
            st.session_state.moved_summary = summary_lines


# ── Affichage des resultats (hors du bloc if uploaded pour persister) ──
if st.session_state.result_excel is not None:
    st.divider()
    delta = st.session_state.score_opt - st.session_state.score_init
    c1, c2 = st.columns(2)
    c1.metric("Score initial",  f"{st.session_state.score_init:.0f}")
    c2.metric("Score optimisé", f"{st.session_state.score_opt:.0f}", delta=f"{delta:+.0f}")
    st.divider()
    st.download_button(
        label="⬇️  Télécharger le fichier résultat (.xlsx)",
        data=st.session_state.result_excel,
        file_name="ville_optimisee.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    with st.expander("📊 Ouvrir dans Google Sheets"):
        st.markdown("""
Le fichier `.xlsx` téléchargé est directement compatible avec Google Sheets, **sans avoir Excel**.

**Sur ordinateur :**
1. Téléchargez le fichier ci-dessus
2. Allez sur [sheets.google.com](https://sheets.google.com) → **Nouveau**
3. **Fichier → Importer** → glissez le `.xlsx` → **Insérer une nouvelle feuille de calcul**

**Sur iPad :**
1. Téléchargez le fichier (il va dans l'app **Fichiers**)
2. Ouvrez l'app **Google Sheets** → **+** → **Importer**
3. Sélectionnez le `.xlsx` dans Fichiers

Les 4 onglets, les couleurs et la mise en forme sont conservés.
        """)

    st.caption(
        "Le fichier contient 4 onglets : "
        "**Liste batiments**, **Synthèse**, **Déplacements**, **Terrain optimisé**."
    )
