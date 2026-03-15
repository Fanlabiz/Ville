import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import time

st.set_page_config(page_title="Placement de Batiments", page_icon="🏙️", layout="wide")
st.title("🏙️ Optimisation de Placement de Batiments")
st.markdown("---")

# ─────────────────────────────────────────────
# LECTURE FICHIER
# ─────────────────────────────────────────────

def lire_terrain(xl):
    df = pd.read_excel(xl, sheet_name=0, header=None)
    terrain = [['0' if pd.isna(v) else str(v).strip() for v in row]
               for _, row in df.iterrows()]
    while terrain and all(v in ('0', '') for v in terrain[-1]):
        terrain.pop()
    cols = max(len(r) for r in terrain)
    return [r + ['0'] * (cols - len(r)) for r in terrain]

def lire_batiments(xl):
    df = pd.read_excel(xl, sheet_name=1, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    batiments = []
    for _, row in df.iterrows():
        nom = str(row.get('Nom', '')).strip()
        if not nom or nom == 'nan':
            continue
        try:
            prod = str(row.get('Production', '')).strip()
            if prod in ('nan', 'Rien', ''):
                prod = ''
            batiments.append({
                'nom': nom,
                'longueur': int(row['Longueur']),
                'largeur': int(row['Largeur']),
                'quantite': int(row['Quantite']),
                'type': str(row.get('Type', 'Neutre')).strip(),
                'culture': float(row['Culture']) if pd.notna(row.get('Culture')) else 0.0,
                'rayonnement': int(row['Rayonnement']) if pd.notna(row.get('Rayonnement')) else 0,
                'boost_25': float(row['Boost 25%']) if pd.notna(row.get('Boost 25%')) else None,
                'boost_50': float(row['Boost 50%']) if pd.notna(row.get('Boost 50%')) else None,
                'boost_100': float(row['Boost 100%']) if pd.notna(row.get('Boost 100%')) else None,
                'production': prod
            })
        except Exception:
            continue
    return batiments

def construire_grille(terrain):
    rows, cols = len(terrain), len(terrain[0])
    grid = np.zeros((rows, cols), dtype=int)
    for r in range(rows):
        for c in range(cols):
            if terrain[r][c] == '1':
                grid[r][c] = 1
    return grid

# ─────────────────────────────────────────────
# UTILITAIRES GRILLE
# ─────────────────────────────────────────────

def peut_placer(grid, r, c, h, w):
    rows, cols = grid.shape
    if r + h > rows or c + w > cols:
        return False
    return np.all(grid[r:r+h, c:c+w] == 1)

def placer_g(grid, r, c, h, w):
    grid[r:r+h, c:c+w] = 0

def liberer_g(grid, r, c, h, w):
    grid[r:r+h, c:c+w] = 1

# ─────────────────────────────────────────────
# CALCUL CULTURE ET BOOST
# ─────────────────────────────────────────────

def zone_rayonnement(r, c, h, w, rayon):
    cases = set()
    for dr in range(-rayon, h + rayon):
        for dc in range(-rayon, w + rayon):
            if dr < 0 or dr >= h or dc < 0 or dc >= w:
                cases.add((r + dr, c + dc))
    return cases

def calc_culture(bat, culturels):
    cases = {(bat['row'] + dr, bat['col'] + dc)
             for dr in range(bat['h']) for dc in range(bat['w'])}
    total = 0.0
    for cult in culturels:
        if cult['rayonnement'] == 0:
            continue
        zone = zone_rayonnement(cult['row'], cult['col'], cult['h'], cult['w'], cult['rayonnement'])
        if cases & zone:
            total += cult['culture']
    return total

def boost_niv(c, b25, b50, b100):
    if b100 is not None and c >= b100:
        return '100%'
    if b50 is not None and c >= b50:
        return '50%'
    if b25 is not None and c >= b25:
        return '25%'
    return '0%'

def score_boost_val(c, b25, b50, b100):
    if b100 is not None and c >= b100:
        return 3
    if b50 is not None and c >= b50:
        return 2
    if b25 is not None and c >= b25:
        return 1
    return 0

def recalc_all(places):
    """Recalcule culture et boost pour tous les producteurs. Retourne le score total."""
    culturels = [b for b in places if b['type'] == 'Culturel' and b['rayonnement'] > 0]
    score = 0
    for b in places:
        if b['type'] == 'Producteur':
            b['culture_recue'] = calc_culture(b, culturels)
            b['boost'] = boost_niv(b['culture_recue'], b['boost_25'], b['boost_50'], b['boost_100'])
            score += score_boost_val(b['culture_recue'], b['boost_25'], b['boost_50'], b['boost_100'])
        else:
            b['culture_recue'] = 0.0
            b['boost'] = 'N/A'
    return score

# ─────────────────────────────────────────────
# FONCTIONS GEOMETRIQUES TERRAIN
# ─────────────────────────────────────────────

def precompute_dist_bord(terrain, rows, cols):
    """Distance de Manhattan au bord le plus proche pour chaque case libre."""
    dist = np.full((rows, cols), 999, dtype=int)
    for r in range(rows):
        for c in range(cols):
            if terrain[r][c] != '1':
                continue
            d = 999
            for dr in range(-20, 21):
                for dc in range(-20, 21):
                    if abs(dr) + abs(dc) >= d:
                        continue
                    rr, cc = r + dr, c + dc
                    if rr < 0 or rr >= rows or cc < 0 or cc >= cols or terrain[rr][cc] == 'X':
                        d = min(d, abs(dr) + abs(dc))
            dist[r][c] = d
    return dist

def dist_bord_bat(r, c, h, w, dist_bord, rows, cols):
    """Distance au bord du coin le plus proche du bâtiment (cases sur terrain)."""
    return min(
        dist_bord[r + dr][c + dc]
        for dr in range(h) for dc in range(w)
        if 0 <= r + dr < rows and 0 <= c + dc < cols
    )

def zone_exploitable_pct(r, c, h, w, rayon, terrain, rows, cols):
    """% de la zone de rayonnement qui tombe sur des cases valides (pas X)."""
    total, sur_terrain = 0, 0
    for dr in range(-rayon, h + rayon):
        for dc in range(-rayon, w + rayon):
            if dr < 0 or dr >= h or dc < 0 or dc >= w:
                total += 1
                rr, cc = r + dr, c + dc
                if 0 <= rr < rows and 0 <= cc < cols and terrain[rr][cc] != 'X':
                    sur_terrain += 1
    return sur_terrain / total if total > 0 else 0

# ─────────────────────────────────────────────
# ALGORITHME D'OPTIMISATION
# ─────────────────────────────────────────────

def optimiser_placement(grid_base, batiments_list, max_time=60, progress_cb=None):
    """
    Placement en 7 passes :
    1.  Placement garanti de TOUS les batiments (tri surface desc, perimetre asc)
    1B. Swaps neutre <-> neutre pour pousser les neutres vers les bords du terrain
    1C. Swaps culturel <-> neutre (objectif combiné) pour placer les culturels
        au centre avec la meilleure zone de rayonnement exploitable
    2.  Swaps producteur <-> neutre pour gagner des boosts
    3.  Swaps producteur <-> producteur pour redistributer les boosts
    4.  Swaps culturel <-> neutre/culturel pour ameliorer les boosts
    5.  Relocation des culturels vers cases libres (objectif combiné boost + culture)
    """
    rows, cols = grid_base.shape
    debut = time.time()

    def elapsed():
        return time.time() - debut

    def remaining():
        return max_time - elapsed()

    # ── Precompute distance au bord pour chaque case ──────────────────────────────
    terrain_ref = None  # sera extrait de grid_base + forme
    dist_bord_arr = None

    # ── PASSE 1 : placement garanti (grand → petit, plus carres en premier) ──────
    if progress_cb:
        progress_cb(0.05, "Passe 1 : placement de tous les batiments...")

    def cle_packing(b):
        s = b['longueur'] * b['largeur']
        p = 2 * (b['longueur'] + b['largeur'])
        return (-s, p)

    tous = sorted(
        [dict(b) for b in batiments_list for _ in range(b['quantite'])],
        key=cle_packing
    )

    grid = grid_base.copy()
    places = []
    non_places_p1 = []

    for bat in tous:
        placed = False
        oris = list({(bat['largeur'], bat['longueur']), (bat['longueur'], bat['largeur'])})
        for h, w in oris:
            if placed:
                break
            for r in range(rows):
                if placed:
                    break
                for c in range(cols):
                    if peut_placer(grid, r, c, h, w):
                        placer_g(grid, r, c, h, w)
                        ori = 'H' if h == bat['largeur'] else 'V'
                        places.append({
                            **bat, 'row': r, 'col': c, 'h': h, 'w': w, 'orientation': ori,
                            'culture_recue': 0.0, 'boost': 'N/A'
                        })
                        placed = True
                        break
        if not placed:
            non_places_p1.append(bat)

    # Backtracking local si necessaire
    if non_places_p1:
        for bat in non_places_p1[:]:
            oris = list({(bat['largeur'], bat['longueur']), (bat['longueur'], bat['largeur'])})
            placed = False
            for target_h, target_w in oris:
                if placed:
                    break
                for j, pb in enumerate(places):
                    if pb['h'] != target_h or pb['w'] != target_w:
                        continue
                    liberer_g(grid, pb['row'], pb['col'], pb['h'], pb['w'])
                    r_new, c_new = None, None
                    for r in range(rows):
                        for c in range(cols):
                            if peut_placer(grid, r, c, target_h, target_w):
                                r_new, c_new = r, c
                                break
                        if r_new is not None:
                            break
                    if r_new is not None and (r_new, c_new) != (pb['row'], pb['col']):
                        placer_g(grid, r_new, c_new, target_h, target_w)
                        r_b, c_b = None, None
                        for r in range(rows):
                            for c in range(cols):
                                if peut_placer(grid, r, c, pb['h'], pb['w']):
                                    r_b, c_b = r, c
                                    break
                            if r_b is not None:
                                break
                        if r_b is not None:
                            placer_g(grid, r_b, c_b, pb['h'], pb['w'])
                            ori = 'H' if target_h == bat['largeur'] else 'V'
                            places.append({
                                **bat, 'row': r_new, 'col': c_new, 'h': target_h, 'w': target_w,
                                'orientation': ori, 'culture_recue': 0.0, 'boost': 'N/A'
                            })
                            places[j] = {**places[j], 'row': r_b, 'col': c_b}
                            non_places_p1.remove(bat)
                            placed = True
                            break
                        else:
                            liberer_g(grid, r_new, c_new, target_h, target_w)
                    if not placed:
                        placer_g(grid, pb['row'], pb['col'], pb['h'], pb['w'])
                    if placed:
                        break

    score_p1 = recalc_all(places)

    if progress_cb:
        progress_cb(0.20, f"Passe 1 : {len(places)} places, score boost={score_p1}")

    if remaining() < 5:
        return places, non_places_p1, int(np.sum(grid == 1))

    # ── Helpers swap ──────────────────────────────────────────────────────────────
    def do_swap(i, j):
        ri, ci2 = places[i]['row'], places[i]['col']
        rj, cj2 = places[j]['row'], places[j]['col']
        places[i]['row'], places[i]['col'] = rj, cj2
        places[j]['row'], places[j]['col'] = ri, ci2
        return ri, ci2, rj, cj2

    def undo_swap(i, j, ri, ci2, rj, cj2):
        places[i]['row'], places[i]['col'] = ri, ci2
        places[j]['row'], places[j]['col'] = rj, cj2

    def objectif_combine():
        culturels_loc = [b for b in places if b['type'] == 'Culturel' and b['rayonnement'] > 0]
        sb, ct = 0, 0.0
        for b in places:
            if b['type'] == 'Producteur' and b['boost_25']:
                c = calc_culture(b, culturels_loc)
                sb += score_boost_val(c, b['boost_25'], b['boost_50'], b['boost_100'])
                ct += c
        return sb * 1_000_000_000 + ct

    def culture_delivree(b):
        zone = zone_rayonnement(b['row'], b['col'], b['h'], b['w'], b['rayonnement'])
        return sum(
            b['culture'] for p in places
            if p['type'] == 'Producteur' and
               {(p['row'] + dr, p['col'] + dc) for dr in range(p['h']) for dc in range(p['w'])} & zone
        )

    # ── PASSE 1B : swap neutre <-> neutre pour pousser les neutres vers les bords ─
    if progress_cb:
        progress_cb(0.25, "Passe 1B : neutres → bordures du terrain...")

    # Precomputer les distances au bord une seule fois
    dist_bord_arr = np.full((rows, cols), 999, dtype=int)
    for r in range(rows):
        for c in range(cols):
            if grid_base[r][c] != 1:
                continue
            d = 999
            for dr in range(-20, 21):
                for dc in range(-20, 21):
                    if abs(dr) + abs(dc) >= d:
                        continue
                    rr, cc = r + dr, c + dc
                    if rr < 0 or rr >= rows or cc < 0 or cc >= cols or grid_base[rr][cc] == 0:
                        d = min(d, abs(dr) + abs(dc))
            dist_bord_arr[r][c] = d

    def dba(r, c, h, w):
        return min(dist_bord_arr[r + dr][c + dc]
                   for dr in range(h) for dc in range(w)
                   if 0 <= r + dr < rows and 0 <= c + dc < cols)

    t1b = time.time()
    MAX_1B = min(remaining() * 0.1, 10)
    ameliorations_1b = 0
    while time.time() - t1b < MAX_1B:
        improved = False
        neutres_idx = [(i, b) for i, b in enumerate(places) if b['type'] == 'Neutre']
        for ni, nb in neutres_idx:
            for nj, mb in neutres_idx:
                if ni >= nj or nb['h'] != mb['h'] or nb['w'] != mb['w']:
                    continue
                d_av = dba(nb['row'], nb['col'], nb['h'], nb['w']) + dba(mb['row'], mb['col'], mb['h'], mb['w'])
                d_ap = dba(mb['row'], mb['col'], nb['h'], nb['w']) + dba(nb['row'], nb['col'], mb['h'], mb['w'])
                if d_ap < d_av:
                    do_swap(ni, nj)
                    ameliorations_1b += 1
                    improved = True
                    break
            if improved:
                break
        if not improved:
            break
    recalc_all(places)

    if progress_cb:
        progress_cb(0.30, f"Passe 1B : {ameliorations_1b} swaps neutres → bords")

    if remaining() < 5:
        return places, non_places_p1, int(np.sum(grid == 1))

    # ── PASSE 1C : swap culturel <-> neutre pour améliorer zone exploitable ───────
    # Utilise objectif_combine pour ne pas degrader les boosts existants
    if progress_cb:
        progress_cb(0.35, "Passe 1C : culturels → meilleures zones de rayonnement...")

    t1c = time.time()
    MAX_1C = min(remaining() * 0.15, 15)
    ameliorations_1c = 0
    s_obj = objectif_combine()

    def score_zone_culturel(r, c, h, w, rayon):
        """Score de qualité de position pour un culturel : zone exploitable + pénalité neutres."""
        if rayon == 0:
            return 0.0
        # Utilise grid_base pour le calcul de l'exploitabilite
        total, sur_terrain = 0, 0
        for dr in range(-rayon, h + rayon):
            for dc in range(-rayon, w + rayon):
                if dr < 0 or dr >= h or dc < 0 or dc >= w:
                    total += 1
                    rr, cc = r + dr, c + dc
                    if 0 <= rr < rows and 0 <= cc < cols and grid_base[rr][cc] == 1:
                        sur_terrain += 1
        expl = sur_terrain / total if total > 0 else 0
        zone = zone_rayonnement(r, c, h, w, rayon)
        non_prod_in_zone = sum(
            1 for p in places if p['type'] in ('Neutre', 'Culturel') and
            any((p['row'] + dr2, p['col'] + dc2) in zone
                for dr2 in range(p['h']) for dc2 in range(p['w']))
        )
        return expl * 10000 - non_prod_in_zone * 100

    while time.time() - t1c < MAX_1C:
        improved = False
        # Trier culturels par culture_delivree croissante (les moins utiles en premier)
        cinfo = sorted(
            [(i, culture_delivree(b)) for i, b in enumerate(places)
             if b['type'] == 'Culturel' and b['rayonnement'] > 0],
            key=lambda x: x[1]
        )
        for ci, _ in cinfo:
            cult = places[ci]
            ch, cw, cr_val = cult['h'], cult['w'], cult['rayonnement']
            candidats = [(j, b) for j, b in enumerate(places)
                         if j != ci and b['type'] == 'Neutre'
                         and b['h'] == ch and b['w'] == cw]
            best_obj, best_j = s_obj, None
            for j, cand in candidats:
                # Vérifie si la position du neutre serait meilleure pour le culturel
                sc_at_cand = score_zone_culturel(cand['row'], cand['col'], ch, cw, cr_val)
                sc_at_cult = score_zone_culturel(cult['row'], cult['col'], ch, cw, cr_val)
                if sc_at_cand > sc_at_cult + 200:  # seuil significatif
                    ri, ci2, rj, cj2 = do_swap(ci, j)
                    obj = objectif_combine()
                    if obj >= best_obj:  # >= pour accepter egalite (améliore zone sans dégrader boosts)
                        best_obj = obj
                        best_j = j
                    undo_swap(ci, j, ri, ci2, rj, cj2)
            if best_j is not None:
                do_swap(ci, best_j)
                s_obj = objectif_combine()
                score_p1 = recalc_all(places)
                ameliorations_1c += 1
                improved = True
                break
        if not improved:
            break

    if progress_cb:
        progress_cb(0.42, f"Passe 1C : {ameliorations_1c} swaps culturels → centre, score={score_p1}")

    if remaining() < 5:
        return places, non_places_p1, int(np.sum(grid == 1))

    # ── PASSE 2 : swaps producteur ↔ neutre (memes dimensions) ─────────────────
    if progress_cb:
        progress_cb(0.48, "Passe 2 : echanges producteur ↔ neutre...")

    t2 = time.time()
    MAX_P2 = min(remaining() * 0.35, 25)
    ameliorations_p2 = 0

    while time.time() - t2 < MAX_P2:
        improved = False
        prod_0 = [(i, b) for i, b in enumerate(places)
                  if b['type'] == 'Producteur' and b['boost'] == '0%' and b['boost_25']]

        for pi, prod in prod_0:
            for ni, neut in enumerate(places):
                if neut['type'] != 'Neutre':
                    continue
                if neut['h'] != prod['h'] or neut['w'] != prod['w']:
                    continue
                ri, ci2, rj, cj2 = do_swap(pi, ni)
                new_score = recalc_all(places)
                if new_score > score_p1:
                    score_p1 = new_score
                    ameliorations_p2 += 1
                    improved = True
                    break
                else:
                    undo_swap(pi, ni, ri, ci2, rj, cj2)
                    recalc_all(places)
            if improved:
                break
        if not improved:
            break

    if progress_cb:
        progress_cb(0.58, f"Passe 2 : {ameliorations_p2} echanges, score={score_p1}")

    if remaining() < 5:
        return places, non_places_p1, int(np.sum(grid == 1))

    # ── PASSE 3 : swaps producteur ↔ producteur ──────────────────────────────────
    if progress_cb:
        progress_cb(0.62, "Passe 3 : echanges producteur ↔ producteur...")

    t3 = time.time()
    MAX_P3 = min(remaining() * 0.3, 15)
    ameliorations_p3 = 0

    while time.time() - t3 < MAX_P3:
        improved = False
        prod_list = [(i, b) for i, b in enumerate(places) if b['type'] == 'Producteur' and b['boost_25']]
        for pi, p1b in prod_list:
            for pj, p2b in prod_list:
                if pi >= pj or p1b['h'] != p2b['h'] or p1b['w'] != p2b['w']:
                    continue
                ri, ci2, rj, cj2 = do_swap(pi, pj)
                ns = recalc_all(places)
                if ns > score_p1:
                    score_p1 = ns; ameliorations_p3 += 1; improved = True; break
                else:
                    undo_swap(pi, pj, ri, ci2, rj, cj2); recalc_all(places)
            if improved:
                break
        if not improved:
            break

    if progress_cb:
        progress_cb(0.70, f"Passe 3 : {ameliorations_p3} echanges, score={score_p1}")

    if remaining() < 5:
        return places, non_places_p1, int(np.sum(grid == 1))

    # ── Objectif enrichi : score_boost * 1e9 + culture_totale + zone_dans_terrain * 0.1 ─
    # Intègre les 3 critères : boosts > culture totale > zones exploitables des culturels
    def objectif_enrichi():
        culturels_loc = [b for b in places if b['type'] == 'Culturel' and b['rayonnement'] > 0]
        sb, ct, zt = 0, 0.0, 0.0
        for b in places:
            if b['type'] == 'Producteur' and b['boost_25']:
                c = calc_culture(b, culturels_loc)
                sb += score_boost_val(c, b['boost_25'], b['boost_50'], b['boost_100'])
                ct += c
            elif b['type'] == 'Culturel' and b['rayonnement'] > 0:
                # Bonus pour les cases de zone dans le terrain (pénalise les bords)
                r0, c0, h0, w0, ray = b['row'], b['col'], b['h'], b['w'], b['rayonnement']
                for dr in range(-ray, h0 + ray):
                    for dc in range(-ray, w0 + ray):
                        if dr < 0 or dr >= h0 or dc < 0 or dc >= w0:
                            rr, cc = r0 + dr, c0 + dc
                            if 0 <= rr < rows and 0 <= cc < cols:
                                zt += 1.0
        return sb * 1_000_000_000 + ct + zt * 0.1

    # ── PASSE 4 : swap culturel ↔ neutre/culturel — objectif enrichi ─────────────
    if progress_cb:
        progress_cb(0.74, "Passe 4 : repositionnement des culturels (swap)...")

    t4 = time.time()
    MAX_P4 = min(remaining() * 0.4, 20)
    ameliorations_p4 = 0
    obj4 = objectif_enrichi()

    while time.time() - t4 < MAX_P4:
        improved = False
        cinfo = sorted(
            [(i, culture_delivree(b)) for i, b in enumerate(places)
             if b['type'] == 'Culturel' and b['rayonnement'] > 0],
            key=lambda x: x[1]
        )
        for ci, _ in cinfo:
            ch, cw = places[ci]['h'], places[ci]['w']
            candidats = [j for j, b in enumerate(places)
                         if j != ci and b['h'] == ch and b['w'] == cw
                         and b['type'] in ('Neutre', 'Culturel')]
            best_obj, best_j = obj4, None
            for j in candidats:
                ri, ci2, rj, cj2 = do_swap(ci, j)
                obj = objectif_enrichi()
                if obj > best_obj:
                    best_obj = obj; best_j = j
                undo_swap(ci, j, ri, ci2, rj, cj2)
            recalc_all(places)
            if best_j is not None:
                do_swap(ci, best_j)
                obj4 = objectif_enrichi()
                score_p1 = recalc_all(places)
                ameliorations_p4 += 1; improved = True; break
        if not improved:
            break

    if progress_cb:
        progress_cb(0.84, f"Passe 4 : {ameliorations_p4} swaps culturels, score={score_p1}")

    if remaining() < 5:
        return places, non_places_p1, int(np.sum(grid == 1))

    # ── PASSE 5 : relocation culturels via objectif combiné ──────────────────────
    if progress_cb:
        progress_cb(0.88, "Passe 5 : relocation des culturels (objectif combiné)...")

    def est_orphelin_fn(b):
        zone = zone_rayonnement(b['row'], b['col'], b['h'], b['w'], b['rayonnement'])
        return not any(
            {(p['row'] + dr, p['col'] + dc)
             for dr in range(p['h']) for dc in range(p['w'])} & zone
            for p in places if p['type'] == 'Producteur'
        )

    t5 = time.time()
    MAX_P5 = min(remaining() * 0.8, 30)
    ameliorations_p5 = 0
    s_obj = objectif_enrichi()

    while time.time() - t5 < MAX_P5:
        improved = False
        cinfo = sorted(
            [(i, culture_delivree(b), est_orphelin_fn(b))
             for i, b in enumerate(places)
             if b['type'] == 'Culturel' and b['rayonnement'] > 0],
            key=lambda x: (0 if x[2] else 1, x[1])
        )
        for ci, _, _ in cinfo:
            b = places[ci]
            h, w = b['h'], b['w']
            liberer_g(grid, b['row'], b['col'], h, w)
            best_obj, best_pos = s_obj, None
            old_r, old_c = b['row'], b['col']
            for r in range(rows):
                for c in range(cols):
                    if peut_placer(grid, r, c, h, w) and (r, c) != (old_r, old_c):
                        places[ci]['row'], places[ci]['col'] = r, c
                        obj = objectif_enrichi()
                        if obj > best_obj:
                            best_obj = obj
                            best_pos = (r, c)
                        places[ci]['row'], places[ci]['col'] = old_r, old_c
            if best_pos:
                r, c = best_pos
                placer_g(grid, r, c, h, w)
                places[ci]['row'], places[ci]['col'] = r, c
                s_obj = objectif_enrichi()
                score_p1 = recalc_all(places)
                ameliorations_p5 += 1
                improved = True
            else:
                placer_g(grid, old_r, old_c, h, w)
            if improved:
                break
        if not improved:
            break

    recalc_all(places)
    if progress_cb:
        progress_cb(0.98, f"Passe 5 : {ameliorations_p5} relocations. Score final={score_p1}")

    return places, non_places_p1, int(np.sum(grid == 1))

# ─────────────────────────────────────────────
# EXPORT EXCEL (5 onglets)
# ─────────────────────────────────────────────

COULEURS = {
    'Culturel':   'FFAA44',
    'Producteur': '66BB66',
    'Neutre':     'BBBBBB',
    'bord':       '444444',
    'libre':      'EAF4EA',
    'occupe':     'DDDDDD',
    'header':     '2B579A',
    'danger':     'C0392B',
    'blanc':      'FFFFFF',
}

def ecrire_resultats(places, non_places, cases_libres, terrain):
    wb = Workbook()

    # ── Onglet 1 : Batiments Places ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Batiments Places"
    hdrs1 = ["Nom", "Type", "Production", "Ligne", "Colonne",
             "Orientation", "Culture Recue", "Boost"]
    for ci, h in enumerate(hdrs1, 1):
        cell = ws1.cell(1, ci, h)
        cell.font = Font(bold=True, color=COULEURS['blanc'])
        cell.fill = PatternFill("solid", fgColor=COULEURS['header'])
        cell.alignment = Alignment(horizontal='center')
    for ri, b in enumerate(places, 2):
        data = [b['nom'], b['type'], b['production'] or '-',
                b['row'] + 1, b['col'] + 1, b['orientation'],
                round(b['culture_recue'], 1), b['boost']]
        fg = COULEURS.get(b['type'], COULEURS['Neutre'])
        for ci, val in enumerate(data, 1):
            cell = ws1.cell(ri, ci, val)
            cell.fill = PatternFill("solid", fgColor=fg)
            cell.alignment = Alignment(horizontal='center')
    for ci in range(1, len(hdrs1) + 1):
        ws1.column_dimensions[get_column_letter(ci)].width = 22
    ws1.freeze_panes = "A2"

    # ── Onglet 2 : Synthese Production ───────────────────────────────────────────
    ws2 = wb.create_sheet("Synthese Production")
    hdrs2 = ["Production", "Nb Batiments", "Culture Totale",
             "Boost 0%", "Boost 25%", "Boost 50%", "Boost 100%"]
    for ci, h in enumerate(hdrs2, 1):
        cell = ws2.cell(1, ci, h)
        cell.font = Font(bold=True, color=COULEURS['blanc'])
        cell.fill = PatternFill("solid", fgColor=COULEURS['header'])
        cell.alignment = Alignment(horizontal='center')
    prod_stats = {}
    for b in places:
        if b['type'] == 'Producteur' and b['production']:
            p = b['production']
            if p not in prod_stats:
                prod_stats[p] = {'nb': 0, 'cult': 0.0, 'b0': 0, 'b25': 0, 'b50': 0, 'b100': 0}
            prod_stats[p]['nb'] += 1
            prod_stats[p]['cult'] += b['culture_recue']
            key = {'0%': 'b0', '25%': 'b25', '50%': 'b50', '100%': 'b100'}.get(b['boost'], 'b0')
            prod_stats[p][key] += 1
    for ri, (p, d) in enumerate(prod_stats.items(), 2):
        ws2.cell(ri, 1, p)
        ws2.cell(ri, 2, d['nb'])
        ws2.cell(ri, 3, round(d['cult'], 1))
        ws2.cell(ri, 4, d['b0'])
        ws2.cell(ri, 5, d['b25'])
        ws2.cell(ri, 6, d['b50'])
        ws2.cell(ri, 7, d['b100'])
    for ci in range(1, 8):
        ws2.column_dimensions[get_column_letter(ci)].width = 20

    # ── Onglet 3 : Terrain visuel ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Terrain")
    rows_t, cols_t = len(terrain), len(terrain[0])

    # Largeur de colonne adaptee pour afficher les noms complets
    COL_W = 7   # largeur en caracteres Excel par case
    ROW_H = 30  # hauteur en points par case (assez pour 2 lignes)

    thin = Side(style='thin', color="999999")
    thick = Side(style='medium', color="333333")

    # 1. Remplir toutes les cases de base (couleur de fond)
    for r in range(rows_t):
        for c in range(cols_t):
            cell = ws3.cell(r + 1, c + 1)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            val = terrain[r][c]
            if val == 'X':
                cell.fill = PatternFill("solid", fgColor=COULEURS['bord'])
                cell.value = 'X'
                cell.font = Font(size=7, color=COULEURS['blanc'])
            elif val == '0':
                cell.fill = PatternFill("solid", fgColor=COULEURS['occupe'])
            else:
                cell.fill = PatternFill("solid", fgColor=COULEURS['libre'])

    # 2. Pour chaque batiment place : fusionner ses cellules, afficher nom + boost
    for b in places:
        r0, c0 = b['row'], b['col']
        h, w = b['h'], b['w']
        fg = COULEURS.get(b['type'], COULEURS['Neutre'])

        # Colorier toutes les cellules du batiment
        for dr in range(h):
            for dc in range(w):
                cell = ws3.cell(r0 + dr + 1, c0 + dc + 1)
                cell.fill = PatternFill("solid", fgColor=fg)

        # Fusionner si le batiment occupe plus d'une cellule
        if h > 1 or w > 1:
            ws3.merge_cells(
                start_row=r0 + 1, start_column=c0 + 1,
                end_row=r0 + h, end_column=c0 + w
            )

        # Construire le texte : nom complet + boost pour les producteurs
        top_cell = ws3.cell(r0 + 1, c0 + 1)
        if b['type'] == 'Producteur' and b['boost'] != 'N/A':
            texte = f"{b['nom']}\n[{b['boost']}]"
        else:
            texte = b['nom']
        top_cell.value = texte

        # Taille de police adaptee a la surface du batiment
        surface = h * w
        if surface >= 20:
            fsize = 8
        elif surface >= 9:
            fsize = 7
        else:
            fsize = 6

        top_cell.font = Font(size=fsize, bold=True)
        top_cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )

        # Bordure epaisse autour du batiment entier
        for dr in range(h):
            for dc in range(w):
                cell = ws3.cell(r0 + dr + 1, c0 + dc + 1)
                left   = thick if dc == 0     else thin
                right  = thick if dc == w - 1 else thin
                top_b  = thick if dr == 0     else thin
                bottom = thick if dr == h - 1 else thin
                cell.border = Border(left=left, right=right, top=top_b, bottom=bottom)

    for ci in range(1, cols_t + 1):
        ws3.column_dimensions[get_column_letter(ci)].width = COL_W
    for ri in range(1, rows_t + 1):
        ws3.row_dimensions[ri].height = ROW_H

    # ── Onglet 4 : Non Places ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Non Places")
    if non_places:
        hdrs4 = ["Nom", "Type", "Longueur", "Largeur", "Cases"]
        for ci, h in enumerate(hdrs4, 1):
            cell = ws4.cell(1, ci, h)
            cell.font = Font(bold=True, color=COULEURS['blanc'])
            cell.fill = PatternFill("solid", fgColor=COULEURS['danger'])
            cell.alignment = Alignment(horizontal='center')
        for ri, b in enumerate(non_places, 2):
            ws4.cell(ri, 1, b['nom'])
            ws4.cell(ri, 2, b['type'])
            ws4.cell(ri, 3, b['longueur'])
            ws4.cell(ri, 4, b['largeur'])
            ws4.cell(ri, 5, b['longueur'] * b['largeur'])
        for ci in range(1, 6):
            ws4.column_dimensions[get_column_letter(ci)].width = 22
    else:
        ws4.cell(1, 1, "Tous les batiments ont ete places !").font = Font(bold=True)

    # ── Onglet 5 : Statistiques ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("Statistiques")
    boosts = {'0%': 0, '25%': 0, '50%': 0, '100%': 0}
    for b in places:
        if b['type'] == 'Producteur' and b['boost'] in boosts:
            boosts[b['boost']] += 1

    stats = [
        ("Batiments places", len(places)),
        ("Batiments non places", len(non_places)),
        ("Cases libres apres placement", cases_libres),
        ("Cases non utilisees (non places)", sum(b['longueur'] * b['largeur'] for b in non_places)),
        ("", ""),
        ("Producteurs boost 0%", boosts['0%']),
        ("Producteurs boost 25%", boosts['25%']),
        ("Producteurs boost 50%", boosts['50%']),
        ("Producteurs boost 100%", boosts['100%']),
        ("Score boost total", sum(v * {'0%': 0, '25%': 1, '50%': 2, '100%': 3}[k]
                                  for k, v in boosts.items())),
    ]
    ws5.cell(1, 1, "Indicateur").font = Font(bold=True)
    ws5.cell(1, 2, "Valeur").font = Font(bold=True)
    for ri, (label, val) in enumerate(stats, 2):
        ws5.cell(ri, 1, label)
        ws5.cell(ri, 2, val)
    ws5.column_dimensions['A'].width = 45
    ws5.column_dimensions['B'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────
# INTERFACE STREAMLIT
# ─────────────────────────────────────────────

uploaded = st.file_uploader("Choisir le fichier Excel d'input (.xlsx)", type=["xlsx"])

if uploaded:
    st.success("Fichier charge avec succes !")

    with st.expander("Apercu du fichier", expanded=False):
        xl = pd.ExcelFile(uploaded)
        for sheet in xl.sheet_names:
            st.subheader(f"Onglet : {sheet}")
            df_prev = pd.read_excel(uploaded, sheet_name=sheet, header=None, nrows=5)
            st.dataframe(df_prev)

    col1, col2 = st.columns(2)
    with col1:
        max_time = st.slider(
            "Temps d'optimisation (secondes)",
            min_value=10, max_value=300, value=60, step=10,
            help="Plus de temps = meilleurs boosts. La passe 1 (placement garanti) prend ~1s."
        )
    with col2:
        st.info("**Strategie :**\n"
                "1. Placement garanti de TOUS les batiments\n"
                "2. Echanges pour maximiser les boosts de production")

    if st.button("Lancer l'optimisation", type="primary"):
        with st.spinner("Lecture du fichier..."):
            terrain = lire_terrain(uploaded)
            batiments_list = lire_batiments(uploaded)
            grid_base = construire_grille(terrain)

        nb_instances = sum(b['quantite'] for b in batiments_list)
        cases_libres_init = int(np.sum(grid_base == 1))
        cases_bats = sum(b['longueur'] * b['largeur'] * b['quantite'] for b in batiments_list)

        c1, c2, c3 = st.columns(3)
        c1.metric("Terrain", f"{len(terrain)} x {len(terrain[0])} cases")
        c2.metric("Cases libres", cases_libres_init)
        c3.metric("Cases batiments", cases_bats)

        if cases_bats > cases_libres_init:
            st.error(f"Impossible de tout placer : {cases_bats} cases de batiments "
                     f"> {cases_libres_init} cases libres. Certains batiments ne seront pas places.")

        progress_bar = st.progress(0, text="Initialisation...")

        def update_progress(pct, msg):
            progress_bar.progress(pct, text=msg)

        places, non_places, cases_libres = optimiser_placement(
            grid_base, batiments_list,
            max_time=max_time,
            progress_cb=update_progress
        )

        progress_bar.progress(1.0, text="Optimisation terminee !")

        st.markdown("---")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Places", f"{len(places)}/{nb_instances}",
                  delta="Tous !" if not non_places else f"-{len(non_places)}")
        m2.metric("Non places", len(non_places))
        m3.metric("Cases libres", cases_libres)
        m4.metric("Cases non places", sum(b['longueur'] * b['largeur'] for b in non_places))

        # Distribution des boosts
        boosts = {'0%': 0, '25%': 0, '50%': 0, '100%': 0}
        for b in places:
            if b['type'] == 'Producteur' and b['boost'] in boosts:
                boosts[b['boost']] += 1
        score_total = sum(v * {'0%': 0, '25%': 1, '50%': 2, '100%': 3}[k] for k, v in boosts.items())

        st.markdown("**Boosts de production obtenus :**")
        bc = st.columns(5)
        bc[0].metric("Boost 0%", boosts['0%'])
        bc[1].metric("Boost 25%", boosts['25%'])
        bc[2].metric("Boost 50%", boosts['50%'])
        bc[3].metric("Boost 100%", boosts['100%'])
        bc[4].metric("Score total", score_total, help="0%=0pt, 25%=1pt, 50%=2pt, 100%=3pt")

        if non_places:
            st.warning(f"**{len(non_places)} batiments non places :**")
            df_np = pd.DataFrame([{
                'Nom': b['nom'], 'Type': b['type'],
                'Taille': f"{b['longueur']}x{b['largeur']}",
                'Cases': b['longueur'] * b['largeur']
            } for b in non_places])
            st.dataframe(df_np, use_container_width=True)
        else:
            st.success("Tous les batiments ont ete places avec succes !")

        with st.spinner("Generation du fichier Excel..."):
            buf = ecrire_resultats(places, non_places, cases_libres, terrain)

        st.download_button(
            label="Telecharger le fichier resultat Excel",
            data=buf,
            file_name="resultats_placement.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
